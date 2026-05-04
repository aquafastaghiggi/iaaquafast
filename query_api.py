"""
================================================================
 AQUAFAST - API local de analise Scanntech
 Expõe consultas deterministicas sobre o DuckDB do projeto
================================================================

 Uso:
   python query_api.py --host 0.0.0.0 --port 8000

Endpoints:
  GET  /health
  POST /ask
  POST /query
================================================================
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import duckdb
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DB_PATH = Path(__file__).with_name("aquafast_scanntech.duckdb")
EXPORT_DIR = Path(__file__).with_name("exports") / "generated"


def normalize(text: str) -> str:
    return " ".join(text.strip().lower().split())


def is_sql_query(text: str) -> bool:
    return bool(re.match(r"^\s*(select|with|show|describe)\b", text, flags=re.IGNORECASE))


def build_query(question: str) -> tuple[str, str]:
    q = normalize(question)

    if any(term in q for term in ["top 20 clientes", "clientes por valor", "ranking clientes", "clientes que mais compraram", "valor total dos clientes"]):
        return "Top 20 clientes por valor total", "SELECT * FROM ranking_clientes LIMIT 20"

    if any(term in q for term in ["top 20 produtos", "produtos mais vendidos", "ranking produtos", "produtos por receita"]):
        return "Top 20 produtos mais vendidos", "SELECT * FROM ranking_produtos LIMIT 20"

    if any(term in q for term in ["vendas por mes", "vendas por mês", "evolucao mensal", "receita por mes", "receita por mês"]):
        return "Vendas por mês", "SELECT * FROM vendas_por_mes ORDER BY mes"

    if any(term in q for term in ["sem compra", "90 dias", "churn"]):
        return (
            "Clientes sem compra há mais de 90 dias",
            """
            SELECT cliente, ultima_compra, total_pedidos, valor_total
            FROM ranking_clientes
            WHERE ultima_compra < CURRENT_DATE - INTERVAL '90 days'
            ORDER BY valor_total DESC
            LIMIT 50
            """.strip(),
        )

    if any(term in q for term in ["1 compra", "uma compra", "apenas uma compra", "risco de churn"]):
        return (
            "Clientes com apenas 1 compra",
            """
            SELECT cliente, total_pedidos, valor_total, primeira_compra, ultima_compra
            FROM ranking_clientes
            WHERE total_pedidos = 1
            ORDER BY valor_total DESC
            LIMIT 50
            """.strip(),
        )

    if any(term in q for term in ["resumo geral", "resumo do arquivo", "visao geral", "visão geral"]):
        return (
            "Resumo geral do arquivo",
            """
            SELECT
                COUNT(*) AS total_registros,
                COUNT(DISTINCT cliente) AS total_clientes,
                ROUND(SUM(valor_total), 2) AS receita_total,
                ROUND(AVG(ticket_medio), 2) AS ticket_medio_geral,
                MIN(primeira_compra) AS periodo_inicio,
                MAX(ultima_compra) AS periodo_fim
            FROM ranking_clientes
            """.strip(),
        )

    if is_sql_query(question):
        return "Consulta SQL livre", question.strip().rstrip(";")

    raise ValueError(
        "Nao identifiquei uma consulta suportada. Tente perguntas como "
        "'Top 20 clientes por valor total', 'Produtos mais vendidos' ou "
        "envie uma consulta SQL SELECT."
    )


def format_markdown(columns: list[str], rows: list[tuple]) -> str:
    if not rows:
        return "_Nenhum resultado encontrado._"

    header = " | ".join(columns)
    separator = " | ".join(["---"] * len(columns))
    lines = [f"| {header} |", f"| {separator} |"]

    for row in rows:
        values = []
        for value in row:
            if value is None:
                values.append("")
            else:
                values.append(str(value))
        lines.append(f"| {' | '.join(values)} |")

    return "\n".join(lines)


def sanitize_filename(text: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", normalize(text)).strip("_")
    return cleaned or "exportacao"


def write_xlsx_report(title: str, columns: list[str], rows: list[tuple]) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{sanitize_filename(title)}_{stamp}.xlsx"
    file_path = EXPORT_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"
    ws.append(columns)
    for row in rows:
        ws.append(list(row))

    for idx, column in enumerate(columns, start=1):
        width = max(len(str(column)), 14)
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, 42)

    wb.save(file_path)
    return file_path


def run_query(sql: str) -> dict:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"Banco nao encontrado: {DB_PATH}")

    con = duckdb.connect(str(DB_PATH), read_only=True)
    try:
        result = con.execute(sql)
        columns = [item[0] for item in result.description]
        rows = result.fetchall()
        return {
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
            "markdown": format_markdown(columns, rows),
        }
    finally:
        con.close()


def export_query(sql: str, title: str) -> dict:
    result = run_query(sql)
    file_path = write_xlsx_report(title, result["columns"], result["rows"])
    return {
        **result,
        "file_name": file_path.name,
        "file_path": str(file_path),
        "download_url": f"http://localhost:8001/download/{file_path.name}",
    }


class Handler(BaseHTTPRequestHandler):
    def _send_json(self, status: int, payload: dict) -> None:
        body = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format: str, *args) -> None:
        return

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path.startswith("/download/"):
            filename = parsed.path.split("/download/", 1)[1]
            file_path = EXPORT_DIR / filename
            if not file_path.exists() or not file_path.is_file():
                self._send_json(404, {"ok": False, "error": "Arquivo nao encontrado"})
                return
            data = file_path.read_bytes()
            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header("Content-Disposition", f'attachment; filename="{file_path.name}"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return

        if parsed.path == "/health":
            try:
                info = run_query("SELECT COUNT(*) AS total_registros FROM scanntech")
                self._send_json(
                    200,
                    {
                        "ok": True,
                        "db_path": str(DB_PATH.name),
                        "total_registros": info["rows"][0][0],
                        "models": ["ranking_clientes", "ranking_produtos", "vendas_por_mes"],
                    },
                )
            except Exception as exc:
                self._send_json(500, {"ok": False, "error": str(exc)})
            return

        if parsed.path == "/":
            self._send_json(200, {"ok": True, "message": "Aquafast Scanntech API"})
            return

        self._send_json(404, {"ok": False, "error": "Endpoint nao encontrado"})

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path not in {"/ask", "/query", "/export"}:
            self._send_json(404, {"ok": False, "error": "Endpoint nao encontrado"})
            return

        try:
            content_length = int(self.headers.get("Content-Length", "0"))
            raw = self.rfile.read(content_length) if content_length else b"{}"
            data = json.loads(raw.decode("utf-8"))
            if parsed.path == "/query":
                sql = str(data.get("sql", "")).strip().rstrip(";")
                title = str(data.get("title", "Consulta SQL")).strip() or "Consulta SQL"
                if not sql:
                    raise ValueError("Campo 'sql' vazio")
            elif parsed.path == "/export":
                sql = str(data.get("sql", "")).strip().rstrip(";")
                title = str(data.get("title", "Exportacao Excel")).strip() or "Exportacao Excel"
                if not sql:
                    raise ValueError("Campo 'sql' vazio")
            else:
                question = str(data.get("question", "")).strip()
                if not question:
                    raise ValueError("Campo 'question' vazio")
                title, sql = build_query(question)

            query_result = export_query(sql, title) if parsed.path == "/export" else run_query(sql)

            self._send_json(
                200,
                {
                    "ok": True,
                    "title": title,
                    "question": data.get("question", ""),
                    "sql": sql,
                    "row_count": query_result["row_count"],
                    "columns": query_result["columns"],
                    "rows": query_result["rows"],
                    "markdown": query_result["markdown"],
                    "file_name": query_result.get("file_name"),
                    "file_path": query_result.get("file_path"),
                    "download_url": query_result.get("download_url"),
                },
            )
        except Exception as exc:
            self._send_json(400, {"ok": False, "error": str(exc)})


def main() -> None:
    parser = argparse.ArgumentParser(description="API local de analise Scanntech")
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--port", type=int, default=8000)
    args = parser.parse_args()

    server = ThreadingHTTPServer((args.host, args.port), Handler)
    print(f"Aquafast Scanntech API ouvindo em http://{args.host}:{args.port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
