"""
Aquafast Scanntech API

FastAPI app for deterministic queries, schema inspection and Excel export.
"""

from __future__ import annotations

import argparse
import os
import math
import json
import re
import threading
import unicodedata
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import duckdb
import mysql.connector
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from uvicorn import run as uvicorn_run

from aquafast_semantics import (
    OFFICIAL_QUESTION_ROUTES,
    list_official_questions,
    normalize_business_question,
    normalize_product_name,
    normalize_volume_signature,
    repair_mojibake,
    resolve_official_route,
    resolve_subgrupo_cigam,
)

APP_NAME = "Aquafast Scanntech API"
DUCKDB_PATH = Path(__file__).with_name("aquafast_scanntech.duckdb")
MYSQL_HOST = os.getenv("AQUAFAST_MYSQL_HOST", "127.0.0.1")
MYSQL_PORT = int(os.getenv("AQUAFAST_MYSQL_PORT", "3306"))
MYSQL_USER = os.getenv("AQUAFAST_MYSQL_USER")
MYSQL_PASSWORD = os.getenv("AQUAFAST_MYSQL_PASSWORD")
MYSQL_DATABASE = os.getenv("AQUAFAST_MYSQL_DATABASE")
CHAT_BACKEND = os.getenv("AQUAFAST_CHAT_BACKEND", "duckdb").strip().lower()
EXPORT_DIR = Path(__file__).with_name("exports") / "generated"
MYSQL_CONNECT_TIMEOUT_SECONDS = int(os.getenv("AQUAFAST_MYSQL_CONNECT_TIMEOUT_SECONDS", "5"))
PORTFOLIO_TABLE = "aquafast_portfolio"
RESOLVED_PRODUCTS_TABLE = "scanntech_produtos_resolvidos"
QUERY_HISTORY_TABLE = "aquafast_query_history"
REPORT_SPECS: dict[str, dict[str, Any]] = {
    "ranking_clientes": {
        "title": "Top clientes Aquafast por caixa",
        "description": "Ranking das lojas/PDVs do portifolio Aquafast por caixas vendidas e receita.",
        "sql": "SELECT * FROM ranking_clientes ORDER BY caixas_vendidas DESC, receita_total DESC, cliente",
    },
    "ranking_produtos": {
        "title": "Top produtos Aquafast por caixa",
        "description": "Ranking dos produtos do portifolio Aquafast por caixas vendidas e receita.",
        "sql": "SELECT * FROM ranking_produtos ORDER BY total_vendas DESC, receita_total DESC, produto",
    },
    "vendas_por_mes": {
        "title": "Vendas Aquafast por mes",
        "description": "Serie mensal do portifolio Aquafast em caixas e receita.",
        "sql": "SELECT * FROM vendas_por_mes ORDER BY mes",
    },
    "vendas_por_cidade": {
        "title": "Vendas Aquafast por cidade",
        "description": "Receita, caixas e unidades por cidade dentro do portfolio Aquafast.",
        "sql": "SELECT * FROM vendas_por_cidade ORDER BY receita_total DESC",
    },
    "market_share_fabricante": {
        "title": "Market share Aquafast por fabricante",
        "description": "Participacao de cada fabricante dentro do mercado Aquafast.",
        "sql": "SELECT * FROM ms_mercado_aquafast ORDER BY total_receita DESC LIMIT 20",
    },
    "concorrentes_por_categoria": {
        "title": "Concorrentes por categoria",
        "description": "Concorrentes que lideram cada categoria no universo Scanntech carregado.",
        "sql": "SELECT * FROM concorrentes_por_categoria ORDER BY categoria, ranking_categoria",
    },
    "share_aquafast_por_categoria": {
        "title": "Share Aquafast por categoria",
        "description": "Participacao da Aquafast versus concorrentes em cada categoria.",
        "sql": "SELECT * FROM share_aquafast_por_categoria ORDER BY share_aquafast_pct DESC, faturamento_total_categoria DESC, categoria",
    },
    "lojas_com_concorrente_sem_aquafast": {
        "title": "Lojas com concorrente sem Aquafast",
        "description": "Lojas que vendem concorrente em categorias onde a Aquafast nao aparece. Usa PDV_ID como chave principal e expõe status_loja quando a ligaÃ§Ã£o nao existir.",
        "sql": "SELECT * FROM lojas_com_concorrente_sem_aquafast ORDER BY faturamento_concorrente DESC, unidades_scanntech DESC, loja, categoria, concorrente",
    },
    "top_concorrentes_por_cidade": {
        "title": "Top concorrentes por cidade",
        "description": "Concorrentes mais fortes por cidade e UF.",
        "sql": "SELECT * FROM top_concorrentes_por_cidade ORDER BY cidade, ranking_cidade",
    },
    "historico_consultas": {
        "title": "Historico de consultas",
        "description": "Ultimas 20 consultas deterministicamente registradas pelo Scanntech Analyst.",
        "sql": "SELECT * FROM aquafast_query_history ORDER BY timestamp DESC, id DESC LIMIT 20",
    },
    "concorrentes_crescimento_90_dias": {
        "title": "Concorrentes em crescimento 90 dias",
        "description": "Concorrentes com maior variacao de faturamento nos 90 dias mais recentes.",
        "sql": "SELECT * FROM concorrentes_crescimento_90_dias ORDER BY variacao_abs DESC, faturamento_90d DESC, concorrente, categoria",
    },
    "vendas_por_estado": {
        "title": "Vendas Aquafast por estado",
        "description": "Receita e cobertura por estado dentro do portfolio Aquafast.",
        "sql": "SELECT * FROM vendas_caixas_estado ORDER BY receita_total DESC",
    },
    "ranking_redes": {
        "title": "Ranking de redes Aquafast",
        "description": "Performance de redes e tipos de loja dentro do portfolio Aquafast.",
        "sql": "SELECT * FROM ranking_redes ORDER BY total_receita DESC",
    },
    "top_produtos_categoria": {
        "title": "Top produtos por categoria Aquafast",
        "description": "Produtos mais fortes por categoria, em caixas, dentro do portfolio Aquafast.",
        "sql": "SELECT * FROM top_produtos_categoria ORDER BY caixas_vendidas DESC, produto_padrao LIMIT 50",
    },
    "auditoria_subgrupo_cigam": {
        "title": "Auditoria SUBGRUPO_CIGAM",
        "description": "Valida a padronizacao dos produtos Aquafast por SUBGRUPO_CIGAM e mostra nomes originais agrupados.",
        "sql": """
            WITH matched AS (
                SELECT
                    p.PROD_NAME AS produto_original_scanntech,
                    ap.SUBGRUPO_CIGAM,
                    ap.SUBGRUPO_LITRAGEM,
                    COUNT(*) AS ocorrencias
                FROM scanntech_produtos_raw p
                JOIN aquafast_portfolio ap
                  ON UPPER(TRIM(COALESCE(p.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE(ap.PROD_CATEGORY, '')))
                 AND REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(UPPER(TRIM(CONCAT(COALESCE(p.PROD_CLASIF_1, ''), ' ', COALESCE(p.PROD_CLASIF_2, '')))), ' ', ''), '.', ''), '/', ''), '-', ''), ',', '')
                   = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(ap.SUBGRUPO_LITRAGEM, ''))), ' ', ''), '.', ''), '/', ''), '-', ''), ',', '')
                WHERE UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) = 'AQUAFAST'
                GROUP BY p.PROD_NAME, ap.SUBGRUPO_CIGAM, ap.SUBGRUPO_LITRAGEM
            ),
            unmatched AS (
                SELECT
                    p.PROD_NAME AS produto_original_scanntech,
                    NULL AS SUBGRUPO_CIGAM,
                    NULL AS SUBGRUPO_LITRAGEM,
                    COUNT(*) AS ocorrencias
                FROM scanntech_produtos_raw p
                LEFT JOIN aquafast_portfolio ap
                  ON UPPER(TRIM(COALESCE(p.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE(ap.PROD_CATEGORY, '')))
                 AND REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(UPPER(TRIM(CONCAT(COALESCE(p.PROD_CLASIF_1, ''), ' ', COALESCE(p.PROD_CLASIF_2, '')))), ' ', ''), '.', ''), '/', ''), '-', ''), ',', '')
                   = REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(ap.SUBGRUPO_LITRAGEM, ''))), ' ', ''), '.', ''), '/', ''), '-', ''), ',', '')
                WHERE UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) = 'AQUAFAST'
                  AND ap.SUBGRUPO_CIGAM IS NULL
                GROUP BY p.PROD_NAME
            )
            SELECT
                produto_original_scanntech,
                SUBGRUPO_CIGAM,
                SUBGRUPO_LITRAGEM,
                ocorrencias
            FROM matched
            UNION ALL
            SELECT
                produto_original_scanntech,
                SUBGRUPO_CIGAM,
                SUBGRUPO_LITRAGEM,
                ocorrencias
            FROM unmatched
            ORDER BY SUBGRUPO_CIGAM IS NULL, ocorrencias DESC, produto_original_scanntech
        """.strip(),
    },
    "auditoria_produtos_sem_subgrupo_cigam": {
        "title": "Auditoria produtos sem SUBGRUPO_CIGAM",
        "description": "Lista produtos Aquafast sem correspondencia no portfolio e sugere SUBGRUPO_CIGAM apenas quando a similaridade e transparente.",
        "sql": "SELECT * FROM auditoria_produtos_sem_subgrupo_cigam ORDER BY faturamento DESC, caixas_vendidas DESC, ocorrencias DESC, produto_original_scanntech",
    },
}
AVAILABLE_REPORTS = list(REPORT_SPECS)
REPORT_PAGE_SIZE_LIMIT = 200
# Limita linhas devolvidas ao chat/Open WebUI (evita travar com SELECT * em fatos enormes).
QUERY_RESULT_ROW_CAP = 2000
# Export Excel: teto para nao estourar RAM com fatos de milhoes de linhas.
EXPORT_RESULT_ROW_CAP = 50_000

_SCHEMA_BOOTSTRAPPED = False
_SCHEMA_BOOTSTRAP_LOCK = threading.Lock()
_QUERY_HISTORY_LOCK = threading.Lock()
AUDIT_PRODUCTS_SENTINEL = "auditoria_produtos_sem_subgrupo_cigam"

if not MYSQL_USER or not MYSQL_PASSWORD or not MYSQL_DATABASE:
    raise RuntimeError(
        "Set AQUAFAST_MYSQL_USER, AQUAFAST_MYSQL_PASSWORD and AQUAFAST_MYSQL_DATABASE "
        "in the environment or .env file."
    )

app = FastAPI(title=APP_NAME, version="1.0.0")


class QuestionRequest(BaseModel):
    question: str = Field(..., min_length=1)


class SQLRequest(BaseModel):
    sql: str = Field(..., min_length=1)
    title: str = Field(default="Consulta SQL")


def normalize(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return " ".join(ascii_text.strip().lower().split())


def _sql_normalized_key(expr: str) -> str:
    """
    Normaliza chaves textuais para join no MySQL/MariaDB sem depender de acento/pontuacao.
    Mantem a regra simples e local para evitar refatoracao ampla.
    """
    return (
        "REPLACE(REPLACE(REPLACE(REPLACE(REPLACE("
        f"UPPER(TRIM(COALESCE({expr}, ''))), ' ', ''), '.', ''), '/', ''), '-', ''), ',', '')"
    )


def _portfolio_join_condition(product_alias: str, portfolio_alias: str) -> str:
    product_group = _sql_normalized_key(
        f"CONCAT(COALESCE({product_alias}.PROD_CLASIF_1, ''), ' ', COALESCE({product_alias}.PROD_CLASIF_2, ''))"
    )
    portfolio_group = _sql_normalized_key(f"{portfolio_alias}.SUBGRUPO_LITRAGEM")
    return (
        f"UPPER(TRIM(COALESCE({product_alias}.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE({portfolio_alias}.PROD_CATEGORY, '')))"
        f" AND {product_group} = {portfolio_group}"
    )


def _audit_normalize(text: str | None) -> str:
    if not text:
        return ""
    repaired = repair_mojibake(str(text))
    normalized = normalize(repaired)
    return normalized.replace("aquafast", " ").replace("produtos", " ").replace("produto", " ")


def _audit_token_set(text: str | None) -> set[str]:
    tokens = re.findall(r"[a-z0-9]+", _audit_normalize(text))
    stopwords = {
        "a",
        "as",
        "de",
        "do",
        "da",
        "dos",
        "das",
        "e",
        "em",
        "com",
        "para",
        "por",
        "um",
        "uma",
        "o",
        "os",
        "na",
        "no",
        "nas",
        "nos",
        "liq",
        "lt",
    }
    return {token for token in tokens if len(token) > 1 and token not in stopwords}


def _audit_size_signature(text: str | None) -> str:
    normalized = _audit_normalize(text)
    normalized = normalized.replace("litros", "l").replace("litro", "l").replace("lts", "l").replace("lt", "l")
    normalized = normalized.replace(",", ".")
    match = re.search(r"\b(\d+(?:\.\d+)?)\s*(ml|l|kg|g)\b", normalized)
    if match:
        return f"{match.group(1)}{match.group(2)}"
    match = re.search(r"\b\d+x(\d+(?:\.\d+)?)\s*(ml|l|kg|g)\b", normalized)
    if match:
        return f"{match.group(1)}{match.group(2)}"
    return ""


def _audit_similarity_score(source_text: str, candidate_text: str) -> float:
    source_tokens = _audit_token_set(source_text)
    candidate_tokens = _audit_token_set(candidate_text)
    if not source_tokens or not candidate_tokens:
        return 0.0
    union = source_tokens | candidate_tokens
    if not union:
        return 0.0
    score = len(source_tokens & candidate_tokens) / len(union)
    source_size = _audit_size_signature(source_text)
    candidate_size = _audit_size_signature(candidate_text)
    if source_size and candidate_size and source_size == candidate_size:
        score += 0.30
    elif source_size and candidate_size and source_size[-1:] == candidate_size[-1:] and source_size[:-1] == candidate_size[:-1]:
        score += 0.15
    return min(score, 1.0)


def _audit_collect_portfolio_candidates(con: Any) -> list[dict[str, Any]]:
    columns, rows = _execute_sql(
        con,
        """
        SELECT PROD_CATEGORY, LITRAGEM, SUBGRUPO_LITRAGEM, QTDE_CX, SUBGRUPO_CIGAM
        FROM aquafast_portfolio
        WHERE SUBGRUPO_CIGAM IS NOT NULL AND TRIM(SUBGRUPO_CIGAM) <> ''
        """,
    )
    return [dict(zip(columns, row)) for row in rows]


def _audit_build_result(con: Any) -> dict[str, Any]:
    columns, rows = _execute_sql(
        con,
        f"""
        SELECT
          r.PROD_ID AS codigo_produto,
          r.produto_original_scanntech AS produto_original_scanntech,
          r.produto_categoria AS produto_categoria,
          r.produto_clasif_1 AS produto_clasif_1,
          r.produto_clasif_2 AS produto_clasif_2,
          COUNT(*) AS ocorrencias,
          ROUND(SUM(s.QTD), 0) AS unidades_scanntech,
          ROUND(SUM(s.VALOR_TOTAL), 2) AS faturamento,
          r.match_mode AS match_mode,
          r.match_confidence AS match_confidence
        FROM scanntech s
        JOIN scanntech_produtos_resolvidos r ON s.COD_PRODUTO = r.PROD_ID
        WHERE r.sem_correspondencia_portfolio = 1
        GROUP BY
          r.PROD_ID,
          r.produto_original_scanntech,
          r.produto_categoria,
          r.produto_clasif_1,
          r.produto_clasif_2,
          r.match_mode,
          r.match_confidence
        ORDER BY faturamento DESC, unidades_scanntech DESC, ocorrencias DESC, produto_original_scanntech
        """,
    )

    portfolio_rows = _audit_collect_portfolio_candidates(con)
    matched_rows: list[tuple[Any, ...]] = []

    for row in rows:
        raw = dict(zip(columns, row))
        raw_category = repair_mojibake(str(raw.get("produto_categoria") or "")).strip()
        raw_text = " ".join(
            part
            for part in [
                raw.get("produto_original_scanntech") or "",
                raw.get("produto_categoria") or "",
                raw.get("produto_clasif_1") or "",
                raw.get("produto_clasif_2") or "",
            ]
            if str(part).strip()
        )
        raw_size = _audit_size_signature(raw_text)

        same_category_candidates = [
            candidate
            for candidate in portfolio_rows
            if _audit_normalize(candidate.get("PROD_CATEGORY")) == _audit_normalize(raw_category)
        ]
        if raw_size:
            size_candidates = [
                candidate
                for candidate in same_category_candidates
                if _audit_size_signature(
                    " ".join(
                        part
                        for part in [
                            candidate.get("PROD_CATEGORY") or "",
                            candidate.get("SUBGRUPO_LITRAGEM") or "",
                            candidate.get("SUBGRUPO_CIGAM") or "",
                            candidate.get("LITRAGEM") or "",
                        ]
                        if str(part).strip()
                    )
                )
                == raw_size
            ]
            same_category_candidates = size_candidates
        candidate_pool = same_category_candidates

        best_candidate: dict[str, Any] | None = None
        best_score = 0.0
        second_score = 0.0

        for candidate in candidate_pool:
            candidate_text = " ".join(
                part
                for part in [
                    candidate.get("PROD_CATEGORY") or "",
                    candidate.get("SUBGRUPO_LITRAGEM") or "",
                    candidate.get("SUBGRUPO_CIGAM") or "",
                    candidate.get("LITRAGEM") or "",
                ]
                if str(part).strip()
            )
            score = _audit_similarity_score(raw_text, candidate_text)
            if score > best_score:
                second_score = best_score
                best_score = score
                best_candidate = candidate
            elif score > second_score:
                second_score = score

        suggested_subgrupo = ""
        suggested_boxes = None
        status = "sem_match"

        if best_candidate is not None and best_score >= 0.30:
            suggested_subgrupo = str(best_candidate.get("SUBGRUPO_CIGAM") or "").strip()
            suggested_qtde = best_candidate.get("QTDE_CX")
            unidades = raw.get("unidades_scanntech")
            if suggested_qtde not in (None, 0) and unidades is not None:
                try:
                    suggested_boxes = _round_half_up(float(unidades) / float(suggested_qtde))
                except Exception:
                    suggested_boxes = None
            if best_score >= 0.58 and (best_score - second_score) >= 0.08:
                status = "revisar_manual"
            else:
                status = "sugestao_baixa_confianca"
        else:
            status = "sem_match"

        matched_rows.append(
            (
                raw.get("codigo_produto"),
                repair_mojibake(str(raw.get("produto_original_scanntech") or "")),
                int(raw.get("ocorrencias") or 0),
                int(suggested_boxes) if suggested_boxes is not None else None,
                int(raw.get("unidades_scanntech") or 0),
                float(raw.get("faturamento") or 0),
                suggested_subgrupo,
                status,
            )
        )

    matched_columns = [
        "codigo_produto",
        "produto_original_scanntech",
        "ocorrencias",
        "caixas_vendidas",
        "unidades_scanntech",
        "faturamento",
        "sugestao_subgrupo_cigam",
        "status",
    ]
    return {
        "columns": matched_columns,
        "rows": matched_rows,
        "row_count": len(matched_rows),
        "markdown": format_markdown(matched_columns, matched_rows),
        "truncated": False,
        "row_cap": None,
    }


def _build_lojas_com_concorrente_sem_aquafast_result() -> dict[str, Any]:
    con = duckdb.connect(str(DUCKDB_PATH), read_only=True)
    try:
        columns, rows = _execute_sql(
            con,
            """
            WITH vendas AS (
              SELECT
                v.PDV_ID,
                COALESCE(NULLIF(d.PDV_NAME, ''), NULLIF(d.PDV_SOCIAL_NAME, ''), CONCAT('PDV ', CAST(v.PDV_ID AS VARCHAR))) AS loja,
                COALESCE(NULLIF(d.PDV_LOCATION, ''), 'SEM CIDADE') AS cidade,
                COALESCE(NULLIF(d.PDV_STATE, ''), 'SEM UF') AS uf,
                CASE WHEN d.PDV_ID IS NULL THEN 'sem_chave_pdv' ELSE 'ok' END AS status_loja,
                p.PROD_CATEGORY AS categoria,
                COALESCE(NULLIF(p.PROD_MANUFACTURER, ''), NULLIF(p.PROD_BRAND, ''), 'SEM FABRICANTE') AS concorrente,
                TRY_CAST(v.SALES_UNITS AS DOUBLE) AS unidades,
                TRY_CAST(v.GROSS_SELLOUT AS DOUBLE) AS receita,
                v.MONTH_ID,
                p.PROD_ID
              FROM scanntech_vendas_raw v
              JOIN scanntech_produtos_raw p
                ON v.PROD_ID = p.PROD_ID
              LEFT JOIN scanntech_clientes_raw d
                ON v.PDV_ID = d.PDV_ID
              WHERE p.PROD_CATEGORY IS NOT NULL
                AND UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) <> 'AQUAFAST'
            ),
            aqua_presence AS (
              SELECT DISTINCT
                v.PDV_ID,
                p.PROD_CATEGORY AS categoria
              FROM scanntech_vendas_raw v
              JOIN scanntech_produtos_raw p
                ON v.PROD_ID = p.PROD_ID
              WHERE p.PROD_CATEGORY IS NOT NULL
                AND UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) = 'AQUAFAST'
                AND v.PDV_ID IS NOT NULL
            )
            SELECT
              loja,
              cidade,
              uf,
              status_loja,
              concorrente,
              v.categoria AS categoria,
              ROUND(SUM(receita), 2) AS faturamento_concorrente,
              ROUND(SUM(unidades), 0) AS unidades_scanntech,
              MAX(MONTH_ID) AS ultima_venda_concorrente
            FROM vendas v
            LEFT JOIN aqua_presence a
              ON a.PDV_ID = v.PDV_ID
             AND a.categoria = v.categoria
            WHERE a.PDV_ID IS NULL
            GROUP BY loja, cidade, uf, status_loja, concorrente, v.categoria
            ORDER BY faturamento_concorrente DESC, unidades_scanntech DESC, loja, categoria, concorrente
            """.strip(),
        )
    finally:
        con.close()

    return {
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
        "markdown": format_markdown(columns, rows),
        "truncated": False,
        "row_cap": None,
    }


def _build_historico_consultas_result() -> dict[str, Any]:
    con = open_connection()
    try:
        columns, rows = _execute_sql(
            con,
            """
            SELECT
              timestamp AS data_hora,
              pergunta,
              report_name AS relatorio,
              metric,
              rows_returned AS linhas_retornadas,
              status
            FROM aquafast_query_history
            ORDER BY timestamp DESC, id DESC
            LIMIT 20
            """.strip(),
        )
    finally:
        con.close()

    return {
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
        "markdown": format_markdown(columns, rows),
        "truncated": False,
        "row_cap": None,
    }


def _is_audit_products_without_subgroup_query(sql: str) -> bool:
    return AUDIT_PRODUCTS_SENTINEL in sql.lower()


def _is_lojas_com_concorrente_sem_aquafast_query(sql: str) -> bool:
    return "lojas_com_concorrente_sem_aquafast" in sql.lower()


def _normalize_sql_for_match(sql: str) -> str:
    return " ".join(sql.strip().rstrip(";").lower().split())


def _infer_report_name_from_sql(sql: str) -> str | None:
    normalized = _normalize_sql_for_match(sql)
    if normalized == _normalize_sql_for_match(
        "SELECT * FROM lojas_com_concorrente_sem_aquafast ORDER BY faturamento_concorrente DESC, unidades_scanntech DESC, loja, categoria, concorrente"
    ):
        return "lojas_com_concorrente_sem_aquafast"
    if normalized == _normalize_sql_for_match(
        "SELECT * FROM aquafast_query_history ORDER BY timestamp DESC, id DESC LIMIT 20"
    ):
        return "historico_consultas"
    if _is_audit_products_without_subgroup_query(sql):
        return "auditoria_produtos_sem_subgrupo_cigam"
    for report_name, spec in REPORT_SPECS.items():
        if report_name in {"lojas_com_concorrente_sem_aquafast", "historico_consultas", "auditoria_produtos_sem_subgrupo_cigam"}:
            continue
        if normalized == _normalize_sql_for_match(str(spec["sql"])):
            return report_name
    return None


def _route_metadata_for_response(question: str, title: str, sql: str) -> dict[str, str]:
    official_route = resolve_official_route(question) or resolve_official_route(title) or resolve_official_route(sql)
    report_name = _infer_report_name_from_sql(sql) or ""
    route = official_route.id if official_route is not None else report_name
    group = repair_mojibake(official_route.title) if official_route is not None else title.strip()
    if not group:
        group = title.strip() or "Consulta SQL"
    intent = route or report_name or ""
    return {
        "route": route,
        "group": group,
        "intent": intent,
        "report_name": report_name or route or "",
    }


def _history_metric_for_report(report_name: str) -> str:
    return {
        "ranking_clientes": "caixas_vendidas",
        "ranking_produtos": "caixas_vendidas",
        "vendas_por_mes": "receita_total",
        "vendas_por_cidade": "receita_total",
        "market_share_fabricante": "market_share_pct",
        "concorrentes_por_categoria": "faturamento",
        "share_aquafast_por_categoria": "share_aquafast_pct",
        "lojas_com_concorrente_sem_aquafast": "faturamento_concorrente",
        "top_concorrentes_por_cidade": "faturamento",
        "concorrentes_crescimento_90_dias": "variacao_abs",
        "vendas_por_estado": "receita_total",
        "ranking_redes": "total_receita",
        "top_produtos_categoria": "caixas_vendidas",
        "auditoria_subgrupo_cigam": "ocorrencias",
        "auditoria_produtos_sem_subgrupo_cigam": "faturamento",
        "historico_consultas": "rows_returned",
    }.get(report_name, "rows_returned")


def _persist_query_history(
    pergunta: str,
    report_name: str,
    metric: str,
    rows_returned: int,
    status: str,
    timestamp_value: datetime | None = None,
) -> dict[str, Any]:
    timestamp = (timestamp_value or datetime.now()).replace(microsecond=0)
    metadata = {
        "history_timestamp": timestamp.isoformat(sep=" "),
        "history_report_name": report_name,
        "history_metric": metric,
        "history_rows_returned": int(rows_returned),
        "history_status": status,
    }
    if report_name == "historico_consultas":
        return metadata

    try:
        with _QUERY_HISTORY_LOCK:
            con = duckdb.connect(str(DUCKDB_PATH))
            try:
                _ensure_query_history_table(con)
                next_id = con.execute(f"SELECT COALESCE(MAX(id), 0) + 1 FROM {QUERY_HISTORY_TABLE}").fetchone()[0]
                con.execute(
                    f"INSERT INTO {QUERY_HISTORY_TABLE} VALUES (?, ?, ?, ?, ?, ?, ?)",
                    [
                        int(next_id or 1),
                        timestamp,
                        pergunta,
                        report_name,
                        metric,
                        int(rows_returned),
                        status,
                    ],
                )
                con.commit()
            finally:
                con.close()
    except Exception as exc:
        print(f"WARN: query history write skipped: {exc}")
    return metadata


def _attach_history_metadata(
    result: dict[str, Any],
    *,
    pergunta: str,
    report_name: str,
    rows_returned: int,
    status: str = "ok",
) -> dict[str, Any]:
    metric = _history_metric_for_report(report_name)
    history = _persist_query_history(pergunta, report_name, metric, rows_returned, status)
    return {**result, **history}


def sanitize_filename(text: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", normalize(text)).strip("_")
    return cleaned or "exportacao"


def format_markdown(columns: list[str], rows: list[tuple[Any, ...]]) -> str:
    if not rows:
        return "_Nenhum resultado encontrado._"

    header = " | ".join(repair_mojibake(str(column)) for column in columns)
    separator = " | ".join(["---"] * len(columns))
    lines = [f"| {header} |", f"| {separator} |"]

    for row in rows:
        values = [_format_ptbr_value(column, value) for column, value in zip(columns, row)]
        lines.append(f"| {' | '.join(values)} |")

    return "\n".join(lines)


def _normalize_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, str):
        return repair_mojibake(value)
    return value


def _format_ptbr_number(value: float | int) -> str:
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return f"{value:,}".replace(",", ".")
    if float(value).is_integer():
        return f"{int(value):,}".replace(",", ".")
    text = f"{float(value):,.2f}"
    return text.replace(",", "X").replace(".", ",").replace("X", ".")


BOX_COUNT_COLUMNS = {
    "caixa",
    "caixas",
    "caixas_vendidas",
    "qtd_caixa",
    "qtd_caixas",
    "total_caixas",
    "quantidade_caixas",
}


def _is_box_count_column(column: str | None) -> bool:
    if not column:
        return False
    normalized = normalize(column)
    if normalized in BOX_COUNT_COLUMNS:
        return True
    if normalized.endswith(" caixas") or normalized.endswith("_caixas"):
        return True
    return False


def _round_half_up(value: float | int | Decimal) -> int:
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _format_ptbr_value(column: str | None, value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        if _is_box_count_column(column):
            return _format_ptbr_number(_round_half_up(value))
        return _format_ptbr_number(value)
    if isinstance(value, Decimal):
        if _is_box_count_column(column):
            return _format_ptbr_number(_round_half_up(value))
        return _format_ptbr_number(value)
    text = repair_mojibake(str(value))
    text = "" if text is None else str(text)
    return "" if text.strip().lower() == "none" else text


def _build_source_note(question: str, title: str, sql: str) -> str:
    text = normalize(" ".join([question, title, sql]))

    if "potencial de venda" in text or "maior potencial" in text:
        return (
            "Fonte: `top_produtos_categoria`. "
            "A consulta usa a presença em PDVs e o volume em caixas como proxy de potencial de venda."
        )
    if any(
        term in text
        for term in [
            "concorrentes por categoria",
            "share aquafast por categoria",
            "lojas com concorrente sem aquafast",
            "top concorrentes por cidade",
            "concorrentes em crescimento 90 dias",
            "concorrentes_por_categoria",
            "share_aquafast_por_categoria",
            "lojas_com_concorrente_sem_aquafast",
            "top_concorrentes_por_cidade",
            "concorrentes_crescimento_90_dias",
        ]
    ):
        return (
            "Fonte: views deterministicas de concorrencia. "
            "A consulta usa o mercado carregado e separa Aquafast de concorrentes sem chamar LLM."
        )
    if any(term in text for term in ["maior concorrente", "concorrentes", "concorrencia", "competidor", "competidores"]):
        return (
            "Fonte: `ms_mercado_aquafast`. "
            "A consulta compara os fabricantes do mercado da categoria e exclui a Aquafast para apontar concorrentes."
        )
    if any(term in text for term in ["market share", "participacao", "share"]):
        return (
            "Fonte: `ms_mercado_aquafast`. "
            "A consulta mede a participacao de cada fabricante dentro do mercado da categoria."
        )
    if any(term in text for term in ["ponto de venda", "pontos de venda", "loja", "lojas", "pdv"]):
        return (
            "Fonte: `ranking_clientes`. "
            "A consulta conta as lojas/PDVs que aparecem com venda Aquafast no periodo carregado."
        )
    if any(term in text for term in ["produto por categoria", "categoria", "litragem", "mix"]):
        return (
            "Fonte: `top_produtos_categoria`. "
            "A consulta cruza o portfolio Aquafast com caixas para enxergar o mix por categoria."
        )
    if any(term in text for term in ["vendas por mes", "vendas por mês", "mensal", "serie mensal", "série mensal"]):
        return (
            "Fonte: `vendas_por_mes`. "
            "A consulta consolida caixas e receita ao longo do tempo para mostrar tendencia mensal."
        )
    if any(term in text for term in ["vendas por estado", "estado", "uf"]):
        return (
            "Fonte: `vendas_caixas_estado`. "
            "A consulta cruza as vendas Aquafast com a UF para mostrar distribuicao geografica."
        )
    if any(term in text for term in ["historico de consultas", "historico consultas", "ultimas consultas", "quais relatorios eu consultei"]):
        return (
            "Fonte: `aquafast_query_history`. "
            "A consulta mostra as 20 consultas deterministicas mais recentes registradas pelo Scanntech Analyst."
        )
    if any(term in text for term in ["top produtos", "ranking produtos", "mais vendidos", "receita por produto", "volume de vendas"]):
        return (
            "Fonte: `ranking_produtos`. "
            "A consulta lista os produtos Aquafast com maior volume em caixas e receita."
        )
    if any(term in text for term in ["clientes", "lojas", "churn", "compra"]):
        return (
            "Fonte: `ranking_clientes`. "
            "A consulta resume as lojas Aquafast por caixas vendidas, receita e recorrencia."
        )
    return "Fonte: consulta local no DuckDB usando as views semanticas da Aquafast."


def _build_source_note_clean(question: str, title: str, sql: str) -> str:
    route = resolve_official_route(question) or resolve_official_route(title) or resolve_official_route(sql)
    if route is not None:
        return repair_mojibake(route.source_note)
    text = normalize_business_question(" ".join([question, title, sql]))
    if "potencial de venda" in text or "maior potencial" in text:
        return repair_mojibake(
            "Fonte: `top_produtos_categoria`. "
            "A consulta usa a presenca em PDVs e o volume em caixas como proxy de potencial de venda."
        )
    if any(
        term in text
        for term in [
            "concorrentes por categoria",
            "share aquafast por categoria",
            "lojas com concorrente sem aquafast",
            "top concorrentes por cidade",
            "concorrentes em crescimento 90 dias",
            "concorrentes_por_categoria",
            "share_aquafast_por_categoria",
            "lojas_com_concorrente_sem_aquafast",
            "top_concorrentes_por_cidade",
            "concorrentes_crescimento_90_dias",
        ]
    ):
        return repair_mojibake(
            "Fonte: views deterministicas de concorrencia. "
            "A consulta usa o mercado carregado e separa Aquafast de concorrentes sem chamar LLM."
        )
    if any(term in text for term in ["maior concorrente", "concorrentes", "concorrencia", "competidor", "competidores"]):
        return repair_mojibake(
            "Fonte: `ms_mercado_aquafast`. "
            "A consulta compara os fabricantes do mercado da categoria e exclui a Aquafast para apontar concorrentes."
        )
    if any(term in text for term in ["market share", "participacao", "share"]):
        return repair_mojibake(
            "Fonte: `ms_mercado_aquafast`. "
            "A consulta mede a participacao de cada fabricante dentro do mercado da categoria."
        )
    if any(term in text for term in ["ponto de venda", "pontos de venda", "loja", "lojas", "pdv"]):
        return repair_mojibake(
            "Fonte: `ranking_clientes`. "
            "A consulta conta as lojas/PDVs que aparecem com venda Aquafast no periodo carregado."
        )
    if any(term in text for term in ["produto por categoria", "categoria", "litragem", "mix"]):
        return repair_mojibake(
            "Fonte: `top_produtos_categoria`. "
            "A consulta cruza o portfolio Aquafast com caixas para enxergar o mix por categoria."
        )
    if any(term in text for term in ["vendas por mes", "mensal", "serie mensal"]):
        return repair_mojibake(
            "Fonte: `vendas_por_mes`. "
            "A consulta consolida caixas e receita ao longo do tempo para mostrar tendencia mensal."
        )
    if any(term in text for term in ["vendas por estado", "estado", "uf"]):
        return repair_mojibake(
            "Fonte: `vendas_caixas_estado`. "
            "A consulta cruza as vendas Aquafast com a UF para mostrar distribuicao geografica."
        )
    if any(term in text for term in ["historico de consultas", "historico consultas", "ultimas consultas", "quais relatorios eu consultei"]):
        return repair_mojibake(
            "Fonte: `aquafast_query_history`. "
            "A consulta mostra as 20 consultas deterministicas mais recentes registradas pelo Scanntech Analyst."
        )
    if any(term in text for term in ["top produtos", "ranking produtos", "mais vendidos", "receita por produto", "volume de vendas"]):
        return repair_mojibake(
            "Fonte: `ranking_produtos`. "
            "A consulta lista os produtos Aquafast com maior volume em caixas e receita."
        )
    if any(term in text for term in ["clientes", "lojas", "churn", "compra"]):
        return repair_mojibake(
            "Fonte: `ranking_clientes`. "
            "A consulta resume as lojas Aquafast por caixas vendidas, receita e recorrencia."
        )
    return repair_mojibake("Fonte: consulta local no DuckDB usando as views semanticas da Aquafast.")


def _duckdb_object_exists(con: duckdb.DuckDBPyConnection, object_name: str) -> bool:
    row = con.execute(
        """
        SELECT 1
        FROM information_schema.tables
        WHERE table_schema NOT IN ('information_schema', 'pg_catalog')
          AND table_name = ?
        LIMIT 1
        """,
        [object_name],
    ).fetchone()
    return row is not None


def _fetch_mysql_portfolio_rows() -> list[tuple[Any, ...]]:
    con = mysql.connector.connect(**_mysql_config())
    cur = None
    try:
        cur = con.cursor()
        cur.execute(
            f"""
            SELECT PROD_CATEGORY, LITRAGEM, SUBGRUPO_LITRAGEM, QTDE_CX, SUBGRUPO_CIGAM
            FROM {PORTFOLIO_TABLE}
            WHERE PROD_CATEGORY IS NOT NULL
            """
        )
        rows = cur.fetchall()
        return rows
    finally:
        if cur is not None:
            try:
                cur.close()
            except Exception:
                pass
        con.close()


def _fetch_resolution_source_rows(con: Any) -> tuple[list[str], list[tuple[Any, ...]], list[str], list[tuple[Any, ...]]]:
    product_columns, product_rows = _execute_sql(
        con,
        """
        SELECT
          PROD_ID,
          PROD_NAME,
          PROD_CATEGORY,
          PROD_CLASIF_1,
          PROD_CLASIF_2,
          PROD_MANUFACTURER,
          PROD_BRAND,
          PROD_NET_WEIGHT
        FROM scanntech_produtos_raw
        WHERE UPPER(TRIM(COALESCE(PROD_MANUFACTURER, ''))) = 'AQUAFAST'
        """.strip(),
    )
    portfolio_columns, portfolio_rows = _execute_sql(
        con,
        """
        SELECT
          PROD_CATEGORY,
          LITRAGEM,
          SUBGRUPO_LITRAGEM,
          QTDE_CX,
          SUBGRUPO_CIGAM
        FROM aquafast_portfolio
        WHERE SUBGRUPO_CIGAM IS NOT NULL AND TRIM(SUBGRUPO_CIGAM) <> ''
        """.strip(),
    )
    return product_columns, product_rows, portfolio_columns, portfolio_rows


def _build_product_resolution_rows(con: Any) -> tuple[list[str], list[tuple[Any, ...]]]:
    product_columns, product_rows, portfolio_columns, portfolio_rows = _fetch_resolution_source_rows(con)
    portfolio = [dict(zip(portfolio_columns, row)) for row in portfolio_rows]

    resolved_columns = [
        "PROD_ID",
        "produto_original_scanntech",
        "produto_categoria",
        "produto_clasif_1",
        "produto_clasif_2",
        "produto_padrao",
        "subgrupo_cigam",
        "subgrupo_litragem",
        "qtde_cx",
        "match_mode",
        "match_confidence",
        "sem_correspondencia_portfolio",
    ]
    resolved_rows: list[tuple[Any, ...]] = []
    for row in product_rows:
        raw = dict(zip(product_columns, row))
        resolved = resolve_subgrupo_cigam(
            raw.get("PROD_NAME"),
            raw.get("PROD_CATEGORY"),
            raw.get("PROD_CLASIF_1"),
            raw.get("PROD_CLASIF_2"),
            portfolio,
        )
        resolved_rows.append(
            (
                raw.get("PROD_ID"),
                repair_mojibake(str(raw.get("PROD_NAME") or "")),
                repair_mojibake(str(raw.get("PROD_CATEGORY") or "")),
                repair_mojibake(str(raw.get("PROD_CLASIF_1") or "")),
                repair_mojibake(str(raw.get("PROD_CLASIF_2") or "")),
                repair_mojibake(str(resolved.get("produto_padrao") or raw.get("PROD_NAME") or "")),
                repair_mojibake(str(resolved.get("subgrupo_cigam") or "")),
                repair_mojibake(str(resolved.get("subgrupo_litragem") or "")),
                int(resolved["qtde_cx"]) if resolved.get("qtde_cx") not in (None, "") else None,
                str(resolved.get("match_mode") or ""),
                str(resolved.get("match_confidence") or ""),
                0 if resolved.get("subgrupo_cigam") else 1,
            )
        )
    return resolved_columns, resolved_rows


def _create_product_resolution_table(con: Any) -> None:
    columns, rows = _build_product_resolution_rows(con)
    con.execute(f"DROP TABLE IF EXISTS {RESOLVED_PRODUCTS_TABLE}")
    con.execute(
        f"""
        CREATE TABLE {RESOLVED_PRODUCTS_TABLE} (
            PROD_ID VARCHAR,
            produto_original_scanntech VARCHAR,
            produto_categoria VARCHAR,
            produto_clasif_1 VARCHAR,
            produto_clasif_2 VARCHAR,
            produto_padrao VARCHAR,
            subgrupo_cigam VARCHAR,
            subgrupo_litragem VARCHAR,
            qtde_cx INTEGER,
            match_mode VARCHAR,
            match_confidence VARCHAR,
            sem_correspondencia_portfolio INTEGER
        )
        """
    )
    if rows:
        placeholders = ", ".join(["?"] * len(columns))
        con.executemany(
            f"INSERT INTO {RESOLVED_PRODUCTS_TABLE} VALUES ({placeholders})",
            rows,
        )


def _ensure_query_history_table(con: Any) -> None:
    con.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {QUERY_HISTORY_TABLE} (
            id BIGINT,
            timestamp TIMESTAMP,
            pergunta VARCHAR,
            report_name VARCHAR,
            metric VARCHAR,
            rows_returned BIGINT,
            status VARCHAR
        )
        """
    )


def _bootstrap_duckdb_aquafast_views() -> None:
    global _SCHEMA_BOOTSTRAPPED
    if _SCHEMA_BOOTSTRAPPED:
        return

    con = duckdb.connect(str(DUCKDB_PATH))
    try:
        if not _duckdb_object_exists(con, PORTFOLIO_TABLE):
            rows = _fetch_mysql_portfolio_rows()
            con.execute(
                f"""
                CREATE TABLE {PORTFOLIO_TABLE} (
                    PROD_CATEGORY VARCHAR,
                    LITRAGEM VARCHAR,
                    SUBGRUPO_LITRAGEM VARCHAR,
                    QTDE_CX INTEGER,
                    SUBGRUPO_CIGAM VARCHAR
                )
                """
            )
            if rows:
                normalized = [
                    (
                        row[0].strip() if row[0] is not None and str(row[0]).strip() else None,
                        row[1].strip() if row[1] is not None and str(row[1]).strip() else None,
                        row[2].strip() if row[2] is not None and str(row[2]).strip() else None,
                        int(row[3]) if row[3] is not None else None,
                        row[4].strip() if row[4] is not None and str(row[4]).strip() else None,
                    )
                    for row in rows
                ]
                con.executemany(
                    f"INSERT INTO {PORTFOLIO_TABLE} VALUES (?, ?, ?, ?, ?)",
                    normalized,
                )
        _create_product_resolution_table(con)
        _ensure_query_history_table(con)

        statements = [
            f"""
            CREATE OR REPLACE VIEW mercado_aquafast AS
            SELECT
              s.MONTH_ID,
              s.QTD AS unidades,
              s.VALOR_TOTAL AS receita,
              p.PROD_ID,
              p.PROD_BARCODE AS ean,
              p.PROD_NAME AS produto,
              p.PROD_MANUFACTURER AS fabricante,
              p.PROD_BRAND AS marca,
              p.PROD_CATEGORY AS categoria,
              p.PROD_CLASIF_1 AS subgrupo_litragem_base,
              p.PROD_NET_WEIGHT AS peso_volume,
              p.PROD_CLASIF_2 AS litragem,
              r.produto_padrao AS produto_padrao,
              NULLIF(r.subgrupo_cigam, '') AS subgrupo_cigam,
              NULLIF(r.subgrupo_litragem, '') AS subgrupo_litragem,
              r.qtde_cx AS unidades_por_caixa,
              r.match_mode AS resolucao_modo,
              r.match_confidence AS resolucao_confianca,
              r.sem_correspondencia_portfolio AS sem_correspondencia_portfolio,
              p.EST_MER_3_DESCRIPTION AS nivel3,
              p.EST_MER_4_DESCRIPTION AS nivel4,
              c.PDV_ID,
              COALESCE(NULLIF(c.PDV_NAME, ''), NULLIF(c.PDV_SOCIAL_NAME, ''), NULLIF(s.RAZAO_SOCIAL, ''), 'SEM LOJA') AS loja,
              COALESCE(NULLIF(c.PDV_LOCATION, ''), 'SEM CIDADE') AS cidade,
              COALESCE(NULLIF(c.PDV_STATE, ''), 'SEM UF') AS estado,
              CASE WHEN c.PDV_ID IS NULL THEN 'sem_chave_pdv' ELSE 'ok' END AS status_loja,
              c.PDV_MICROREGION AS microrregiao,
              c.PDV_STORE_CHAIN AS rede,
              c.STORE_CLASSIFICATION AS tipo_loja,
              c.PDV_CHECKOUTS AS caixas,
              c.PDV_CNPJ AS cnpj_loja,
              c.PDV_SOCIAL_NAME AS razao_social_loja,
              CASE WHEN LOWER(COALESCE(p.PROD_MANUFACTURER, '')) = 'aquafast' THEN 1 ELSE 0 END AS is_aquafast
            FROM scanntech s
            LEFT JOIN scanntech_produtos_raw p ON s.COD_PRODUTO = p.PROD_ID
            LEFT JOIN scanntech_produtos_resolvidos r ON p.PROD_ID = r.PROD_ID
            LEFT JOIN scanntech_clientes_raw c
              ON TRY_CAST(s.CNPJ AS BIGINT) = c.PDV_ID
            WHERE p.PROD_CATEGORY IN (SELECT DISTINCT PROD_CATEGORY FROM aquafast_portfolio)
            """,
            f"""
            CREATE OR REPLACE VIEW vendas_em_caixas AS
            SELECT
              m.MONTH_ID,
              m.fabricante,
              m.marca,
              m.categoria,
              m.subgrupo_litragem_base,
              m.litragem,
              m.produto,
              m.estado,
              m.microrregiao,
              m.rede,
              m.tipo_loja,
              m.loja,
              m.PDV_ID,
              m.is_aquafast,
              ROUND(SUM(m.unidades), 0) AS unidades_scanntech,
              ROUND(SUM(m.unidades), 0) AS total_unidades,
              ROUND(SUM(m.unidades), 0) AS unidades,
              ROUND(SUM(m.receita), 2) AS total_receita,
              ROUND(SUM(m.receita), 2) AS receita,
              m.unidades_por_caixa,
              ROUND(SUM(m.unidades) / NULLIF(m.unidades_por_caixa, 0), 0) AS total_caixas,
              ROUND(SUM(m.unidades) / NULLIF(m.unidades_por_caixa, 0), 0) AS caixas,
              m.subgrupo_litragem AS subgrupo_litragem,
              NULLIF(m.subgrupo_cigam, '') AS subgrupo_cigam,
              COALESCE(NULLIF(m.subgrupo_cigam, ''), m.produto_padrao, m.produto) AS produto_padrao,
              m.sem_correspondencia_portfolio AS sem_correspondencia_portfolio,
              ROUND(SUM(m.receita) / NULLIF(SUM(m.unidades), 0), 2) AS preco_medio_unitario,
              ROUND(SUM(m.receita) / NULLIF(SUM(m.unidades) / NULLIF(m.unidades_por_caixa, 0), 0), 2) AS preco_medio_caixa
            FROM mercado_aquafast m
            GROUP BY
              m.MONTH_ID, m.fabricante, m.marca, m.categoria, m.subgrupo_litragem_base, m.litragem, m.produto,
              m.estado, m.microrregiao, m.rede, m.tipo_loja, m.loja, m.PDV_ID,
              m.is_aquafast, m.unidades_por_caixa, m.subgrupo_litragem, m.subgrupo_cigam, m.produto_padrao, m.sem_correspondencia_portfolio
            """,
            """
            CREATE OR REPLACE VIEW ranking_clientes AS
            SELECT
              loja AS cliente,
              ROUND(SUM(unidades), 0) AS unidades_scanntech,
              ROUND(SUM(caixas), 0) AS caixas_vendidas,
              ROUND(SUM(receita), 2) AS receita_total,
              ROUND(SUM(receita) / NULLIF(SUM(caixas), 0), 2) AS ticket_medio_caixa,
              MIN(MONTH_ID) AS primeira_compra,
              MAX(MONTH_ID) AS ultima_compra
            FROM vendas_em_caixas
            WHERE is_aquafast = 1
            GROUP BY loja
            """,
            """
            CREATE OR REPLACE VIEW ranking_produtos AS
            SELECT
              produto_padrao,
              MIN(produto) AS produto_original_exemplo,
              subgrupo_cigam,
              subgrupo_litragem,
              categoria,
              litragem,
              fabricante,
              marca,
              COUNT(DISTINCT produto) AS variacoes_produto_original,
              ROUND(SUM(unidades), 0) AS unidades_scanntech,
              ROUND(SUM(unidades), 0) AS total_unidades,
              ROUND(SUM(unidades), 0) AS total_vendas,
              ROUND(SUM(unidades) / NULLIF(MAX(unidades_por_caixa), 0), 0) AS caixas_vendidas,
              ROUND(SUM(receita), 2) AS receita_total,
              ROUND(SUM(receita) / NULLIF(SUM(unidades) / NULLIF(MAX(unidades_por_caixa), 0), 0), 2) AS preco_medio_caixa
            FROM vendas_em_caixas
            WHERE is_aquafast = 1
            GROUP BY produto_padrao, subgrupo_cigam, subgrupo_litragem, categoria, litragem, fabricante, marca
            """,
            """
            CREATE OR REPLACE VIEW vendas_por_mes AS
            SELECT
              SUBSTR(CAST(MONTH_ID AS VARCHAR), 1, 4) || '-' || SUBSTR(CAST(MONTH_ID AS VARCHAR), 5, 2) AS mes,
              ROUND(SUM(unidades), 0) AS unidades_scanntech,
              ROUND(SUM(caixas), 0) AS caixas_vendidas,
              ROUND(SUM(receita), 2) AS receita_total,
              COUNT(DISTINCT PDV_ID) AS pdvs_ativos,
              COUNT(DISTINCT categoria) AS categorias
            FROM vendas_em_caixas
            WHERE is_aquafast = 1
            GROUP BY 1
            """,
            """
            CREATE OR REPLACE VIEW ms_mercado_aquafast AS
            SELECT
              fabricante,
              COUNT(DISTINCT produto) AS skus,
              COUNT(DISTINCT PDV_ID) AS pdvs,
              ROUND(SUM(unidades), 0) AS total_unidades,
              ROUND(SUM(receita), 2) AS total_receita,
              ROUND(SUM(receita) / NULLIF((SELECT SUM(receita) FROM mercado_aquafast), 0) * 100, 2) AS market_share_pct,
              MAX(is_aquafast) AS is_aquafast
            FROM mercado_aquafast
            WHERE fabricante IS NOT NULL
            GROUP BY fabricante
            ORDER BY total_receita DESC
            """,
            """
            CREATE OR REPLACE VIEW concorrentes_por_categoria AS
            WITH totais AS (
              SELECT
                categoria,
                ROUND(SUM(receita), 2) AS faturamento_total_categoria
              FROM mercado_aquafast
              WHERE categoria IS NOT NULL
              GROUP BY categoria
            ),
            base AS (
              SELECT
                categoria,
                COALESCE(NULLIF(fabricante, ''), NULLIF(marca, ''), 'SEM FABRICANTE') AS concorrente,
                ROUND(SUM(receita), 2) AS faturamento,
                ROUND(SUM(unidades), 0) AS unidades_scanntech,
                ROUND(SUM(caixas), 0) AS caixas_vendidas
              FROM mercado_aquafast
              WHERE categoria IS NOT NULL
                AND is_aquafast = 0
              GROUP BY categoria, COALESCE(NULLIF(fabricante, ''), NULLIF(marca, ''), 'SEM FABRICANTE')
            )
            SELECT
              base.categoria,
              base.concorrente,
              base.faturamento,
              base.unidades_scanntech,
              base.caixas_vendidas,
              ROUND(base.faturamento / NULLIF(totais.faturamento_total_categoria, 0) * 100, 2) AS participacao_categoria_pct,
              ROW_NUMBER() OVER (
                PARTITION BY base.categoria
                ORDER BY base.faturamento DESC, base.unidades_scanntech DESC, base.concorrente
              ) AS ranking_categoria
            FROM base
            JOIN totais ON totais.categoria = base.categoria
            ORDER BY base.categoria, ranking_categoria, base.concorrente
            """,
            """
            CREATE OR REPLACE VIEW share_aquafast_por_categoria AS
            WITH base AS (
              SELECT
                categoria,
                ROUND(SUM(CASE WHEN is_aquafast = 1 THEN receita ELSE 0 END), 2) AS faturamento_aquafast,
                ROUND(SUM(CASE WHEN is_aquafast = 0 THEN receita ELSE 0 END), 2) AS faturamento_concorrentes,
                ROUND(SUM(receita), 2) AS faturamento_total_categoria,
                ROUND(SUM(CASE WHEN is_aquafast = 1 THEN unidades ELSE 0 END), 0) AS unidades_aquafast,
                ROUND(SUM(CASE WHEN is_aquafast = 0 THEN unidades ELSE 0 END), 0) AS unidades_concorrentes
              FROM mercado_aquafast
              WHERE categoria IS NOT NULL
              GROUP BY categoria
            )
            SELECT
              categoria,
              faturamento_total_categoria,
              faturamento_aquafast,
              faturamento_concorrentes,
              ROUND(faturamento_aquafast / NULLIF(faturamento_total_categoria, 0) * 100, 2) AS share_aquafast_pct,
              unidades_aquafast,
              unidades_concorrentes
            FROM base
            ORDER BY share_aquafast_pct DESC, faturamento_total_categoria DESC, categoria
            """,
            """
            CREATE OR REPLACE VIEW top_concorrentes_por_cidade AS
            WITH base AS (
              SELECT
                COALESCE(NULLIF(cidade, ''), 'SEM CIDADE') AS cidade,
                COALESCE(NULLIF(estado, ''), 'SEM UF') AS uf,
                COALESCE(NULLIF(fabricante, ''), NULLIF(marca, ''), 'SEM FABRICANTE') AS concorrente,
                ROUND(SUM(receita), 2) AS faturamento,
                ROUND(SUM(unidades), 0) AS unidades_scanntech
              FROM mercado_aquafast
              WHERE is_aquafast = 0
                AND cidade IS NOT NULL
              GROUP BY
                COALESCE(NULLIF(cidade, ''), 'SEM CIDADE'),
                COALESCE(NULLIF(estado, ''), 'SEM UF'),
                COALESCE(NULLIF(fabricante, ''), NULLIF(marca, ''), 'SEM FABRICANTE')
            ),
            ranked AS (
              SELECT
                cidade,
                uf,
                concorrente,
                faturamento,
                unidades_scanntech,
                ROW_NUMBER() OVER (
                  PARTITION BY cidade, uf
                  ORDER BY faturamento DESC, unidades_scanntech DESC, concorrente
                ) AS ranking_cidade
              FROM base
            )
            SELECT
              cidade,
              uf,
              concorrente,
              faturamento,
              unidades_scanntech,
              ranking_cidade
            FROM ranked
            ORDER BY cidade, ranking_cidade, concorrente
            """,
            """
            CREATE OR REPLACE VIEW concorrentes_crescimento_90_dias AS
            WITH meses AS (
              SELECT DISTINCT MONTH_ID
              FROM mercado_aquafast
              WHERE MONTH_ID IS NOT NULL
              ORDER BY MONTH_ID DESC
              LIMIT 6
            ),
            base AS (
              SELECT
                COALESCE(NULLIF(m.fabricante, ''), NULLIF(m.marca, ''), 'SEM FABRICANTE') AS concorrente,
                m.categoria,
                m.unidades,
                m.receita,
                DENSE_RANK() OVER (ORDER BY m.MONTH_ID DESC) AS month_rank
              FROM mercado_aquafast m
              JOIN meses x ON x.MONTH_ID = m.MONTH_ID
              WHERE m.is_aquafast = 0
                AND m.categoria IS NOT NULL
            ),
            agg AS (
              SELECT
                concorrente,
                categoria,
                ROUND(SUM(CASE WHEN month_rank <= 3 THEN receita ELSE 0 END), 2) AS faturamento_90d,
                ROUND(SUM(CASE WHEN month_rank BETWEEN 4 AND 6 THEN receita ELSE 0 END), 2) AS faturamento_90d_anterior,
                ROUND(SUM(CASE WHEN month_rank <= 3 THEN unidades ELSE 0 END), 0) AS unidades_90d,
                ROUND(SUM(CASE WHEN month_rank BETWEEN 4 AND 6 THEN unidades ELSE 0 END), 0) AS unidades_90d_anterior
              FROM base
              GROUP BY concorrente, categoria
            )
            SELECT
              concorrente,
              categoria,
              faturamento_90d,
              faturamento_90d_anterior,
              ROUND(faturamento_90d - faturamento_90d_anterior, 2) AS variacao_abs,
              ROUND((faturamento_90d - faturamento_90d_anterior) / NULLIF(faturamento_90d_anterior, 0) * 100, 2) AS variacao_pct,
              unidades_90d,
              unidades_90d_anterior
            FROM agg
            ORDER BY variacao_abs DESC, faturamento_90d DESC, concorrente, categoria
            """,
            """
            CREATE OR REPLACE VIEW concorrencia_por_categoria AS
            SELECT
              categoria,
              litragem,
              fabricante,
              ROUND(SUM(unidades), 0) AS unidades,
              ROUND(SUM(receita), 2) AS receita,
              COUNT(DISTINCT PDV_ID) AS pdvs,
              MAX(is_aquafast) AS is_aquafast
            FROM mercado_aquafast
            WHERE categoria IS NOT NULL
            GROUP BY categoria, litragem, fabricante
            ORDER BY categoria, litragem, receita DESC
            """,
            """
            CREATE OR REPLACE VIEW comparativo_preco AS
            SELECT
              categoria,
              litragem,
              fabricante,
              COUNT(DISTINCT PDV_ID) AS pdvs,
              ROUND(SUM(unidades), 0) AS unidades,
              ROUND(SUM(receita) / NULLIF(SUM(unidades), 0), 2) AS preco_medio_unitario,
              MAX(is_aquafast) AS is_aquafast
            FROM mercado_aquafast
            WHERE litragem IS NOT NULL AND fabricante IS NOT NULL
            GROUP BY categoria, litragem, fabricante
            ORDER BY categoria, litragem, preco_medio_unitario
            """,
            """
            CREATE OR REPLACE VIEW vendas_caixas_estado AS
            SELECT
              estado,
              categoria,
              litragem,
              fabricante,
              is_aquafast,
              ROUND(SUM(unidades), 0) AS unidades_scanntech,
              ROUND(SUM(caixas), 0) AS caixas_vendidas,
              ROUND(SUM(receita), 2) AS receita_total,
              COUNT(DISTINCT PDV_ID) AS pdvs
            FROM vendas_em_caixas
            WHERE is_aquafast = 1
            GROUP BY estado, categoria, litragem, fabricante, is_aquafast
            ORDER BY estado, categoria, caixas_vendidas DESC
            """,
            """
            CREATE OR REPLACE VIEW vendas_por_cidade AS
            SELECT
              COALESCE(m.cidade, 'SEM CIDADE') AS cidade,
              COALESCE(m.estado, 'SEM UF') AS estado,
              COUNT(DISTINCT m.PDV_ID) AS pdvs,
              ROUND(SUM(m.unidades), 0) AS unidades_scanntech,
              ROUND(SUM(m.unidades / NULLIF(m.unidades_por_caixa, 0)), 0) AS caixas_vendidas,
              ROUND(SUM(m.receita), 2) AS receita_total
            FROM mercado_aquafast m
            WHERE m.is_aquafast = 1
            GROUP BY COALESCE(m.cidade, 'SEM CIDADE'), COALESCE(m.estado, 'SEM UF')
            ORDER BY receita_total DESC, caixas_vendidas DESC, cidade
            """,
            """
            CREATE OR REPLACE VIEW resumo_mercado_aquafast AS
            SELECT
              MONTH_ID AS mes,
              COUNT(DISTINCT categoria) AS categorias_monitoradas,
              COUNT(DISTINCT fabricante) AS fabricantes_no_mercado,
              COUNT(DISTINCT CASE WHEN is_aquafast = 1 THEN PDV_ID END) AS pdvs_com_aquafast,
              COUNT(DISTINCT CASE WHEN is_aquafast = 0 THEN PDV_ID END) AS pdvs_so_concorrencia,
              ROUND(SUM(CASE WHEN is_aquafast = 1 THEN receita ELSE 0 END), 2) AS receita_aquafast,
              ROUND(SUM(CASE WHEN is_aquafast = 0 THEN receita ELSE 0 END), 2) AS receita_concorrencia,
              ROUND(SUM(receita), 2) AS receita_total_mercado,
              ROUND(SUM(CASE WHEN is_aquafast = 1 THEN receita ELSE 0 END) / NULLIF(SUM(receita), 0) * 100, 2) AS share_aquafast_pct
            FROM mercado_aquafast
            GROUP BY MONTH_ID
            """,
            f"""
            CREATE OR REPLACE VIEW top_produtos_categoria AS
            SELECT
              m.categoria,
              COALESCE(NULLIF(m.subgrupo_cigam, ''), m.produto_padrao, m.produto) AS produto_padrao,
              MIN(m.produto) AS produto_original_exemplo,
              NULLIF(m.subgrupo_cigam, '') AS subgrupo_cigam,
              m.subgrupo_litragem AS subgrupo_litragem,
              m.fabricante,
              m.marca,
              COUNT(DISTINCT m.produto) AS variacoes_produto_original,
              ROUND(SUM(m.unidades) / NULLIF(MAX(m.unidades_por_caixa), 0), 0) AS caixas_vendidas,
              ROUND(SUM(m.unidades), 0) AS unidades_scanntech,
              ROUND(SUM(m.unidades), 0) AS total_unidades,
              ROUND(SUM(m.receita), 2) AS total_receita,
              ROUND(SUM(m.receita) / NULLIF(SUM(m.unidades) / NULLIF(MAX(m.unidades_por_caixa), 0), 0), 2) AS preco_medio_caixa,
              COUNT(DISTINCT m.PDV_ID) AS pdvs_com_venda
            FROM mercado_aquafast m
            WHERE m.is_aquafast = 1
            GROUP BY m.categoria, COALESCE(NULLIF(m.subgrupo_cigam, ''), m.produto_padrao, m.produto), NULLIF(m.subgrupo_cigam, ''), m.subgrupo_litragem, m.fabricante, m.marca
            ORDER BY caixas_vendidas DESC, total_receita DESC, produto_padrao
            """,
            """
            CREATE OR REPLACE VIEW ranking_redes AS
            SELECT
              COALESCE(rede, 'SEM REDE') AS rede,
              tipo_loja,
              COUNT(DISTINCT PDV_ID) AS total_lojas,
              COUNT(DISTINCT fabricante) AS fabricantes,
              ROUND(SUM(caixas), 0) AS caixas_vendidas,
              ROUND(SUM(receita), 2) AS total_receita
            FROM vendas_em_caixas
            WHERE is_aquafast = 1
            GROUP BY COALESCE(rede, 'SEM REDE'), tipo_loja
            ORDER BY total_receita DESC
            """,
        ]
        for statement in statements:
            con.execute(statement)
        con.commit()
        _SCHEMA_BOOTSTRAPPED = True
    finally:
        con.close()


def _mysql_config() -> dict[str, Any]:
    return {
        "host": MYSQL_HOST,
        "port": MYSQL_PORT,
        "user": MYSQL_USER,
        "password": MYSQL_PASSWORD,
        "database": MYSQL_DATABASE,
        "connection_timeout": MYSQL_CONNECT_TIMEOUT_SECONDS,
    }


def _bootstrap_compatibility_views() -> None:
    global _SCHEMA_BOOTSTRAPPED
    if _SCHEMA_BOOTSTRAPPED:
        return

    con = mysql.connector.connect(**_mysql_config())
    try:
        cur = con.cursor()
        statements = [
            """
            CREATE OR REPLACE VIEW scanntech_clientes_raw AS
            SELECT
              PDV_ID,
              PDV_CODE,
              PDV_NAME,
              PDV_ADDRESS,
              PDV_LOCATION,
              PDV_STATE,
              CAST(PDV_CHECKOUTS AS UNSIGNED) AS PDV_CHECKOUTS,
              PDV_CLASIF_1,
              PDV_CLASIF_2,
              PDV_CLASIF_3,
              PDV_CLASIF_4,
              PDV_CLASIF_5,
              PDV_CNPJ,
              PDV_SOCIAL_NAME,
              PDV_STORE_CHAIN,
              STORE_CLASSIFICATION,
              PDV_MICROREGION
            FROM pdv
            """,
            """
            CREATE OR REPLACE VIEW scanntech_produtos_raw AS
            SELECT
              PROD_ID,
              PROD_BARCODE,
              PROD_NAME,
              PROD_MANUFACTURER,
              PROD_BRAND,
              PROD_CATEGORY,
              PROD_NET_WEIGHT,
              PROD_CLASIF_1,
              PROD_CLASIF_2,
              PROD_CLASIF_3,
              PROD_CLASIF_4,
              PROD_CLASIF_5,
              EST_MER_1_DESCRIPTION,
              EST_MER_2_DESCRIPTION,
              EST_MER_3_DESCRIPTION,
              EST_MER_4_DESCRIPTION,
              EST_MER_5_DESCRIPTION,
              EST_MER_ID,
              PACK_QUANTITY,
              CONTENT_BARCODE
            FROM prd
            """,
            """
            CREATE OR REPLACE VIEW scanntech_vendas_raw AS
            SELECT
              MONTH_ID,
              PDV_ID,
              PROD_ID,
              SALES_UNITS,
              GROSS_SELLOUT
            FROM vta
            """,
            """
            CREATE OR REPLACE VIEW scanntech AS
            SELECT
              v.MONTH_ID,
              p.PDV_CNPJ AS CNPJ,
              COALESCE(p.PDV_SOCIAL_NAME, p.PDV_NAME) AS RAZAO_SOCIAL,
              v.PROD_ID AS COD_PRODUTO,
              pr.PROD_NAME AS DESC_PRODUTO,
              v.SALES_UNITS AS QTD,
              v.GROSS_SELLOUT AS VALOR_TOTAL,
              ROUND(v.GROSS_SELLOUT / NULLIF(v.SALES_UNITS, 0), 5) AS VALOR_UNITARIO,
              STR_TO_DATE(CONCAT(v.MONTH_ID, '01'), '%Y%m%d') AS DATA_VENDA
            FROM vta v
            LEFT JOIN prd pr ON v.PROD_ID = pr.PROD_ID
            LEFT JOIN pdv p ON v.PDV_ID = p.PDV_ID
            """,
            """
            DROP VIEW IF EXISTS ranking_clientes
            """,
            """
            DROP TABLE IF EXISTS ranking_clientes
            """,
            """
            CREATE TABLE ranking_clientes AS
            SELECT
              COALESCE(p.PDV_SOCIAL_NAME, p.PDV_NAME) AS cliente,
              ROUND(SUM(v.SALES_UNITS), 0) AS unidades_scanntech,
              ROUND(SUM(v.SALES_UNITS / NULLIF(ap.QTDE_CX, 0)), 0) AS caixas_vendidas,
              ROUND(SUM(v.GROSS_SELLOUT), 2) AS receita_total,
              ROUND(SUM(v.GROSS_SELLOUT) / NULLIF(SUM(v.SALES_UNITS / NULLIF(ap.QTDE_CX, 0)), 0), 2) AS ticket_medio_caixa,
              MIN(STR_TO_DATE(CONCAT(v.MONTH_ID, '01'), '%Y%m%d')) AS primeira_compra,
              MAX(STR_TO_DATE(CONCAT(v.MONTH_ID, '01'), '%Y%m%d')) AS ultima_compra
            FROM vta v
            JOIN pdv p ON v.PDV_ID = p.PDV_ID
            JOIN prd pr ON v.PROD_ID = pr.PROD_ID
            LEFT JOIN {PORTFOLIO_TABLE} ap
              ON UPPER(TRIM(COALESCE(pr.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE(ap.PROD_CATEGORY, '')))
             AND UPPER(TRIM(COALESCE(pr.PROD_CLASIF_2, ''))) = UPPER(TRIM(COALESCE(ap.LITRAGEM, '')))
            WHERE pr.PROD_CATEGORY IS NOT NULL
            GROUP BY p.PDV_ID, p.PDV_SOCIAL_NAME, p.PDV_NAME
            """,
            """
            DROP VIEW IF EXISTS ranking_produtos
            """,
            """
            DROP TABLE IF EXISTS ranking_produtos
            """,
            """
            CREATE TABLE ranking_produtos AS
            SELECT
              pr.PROD_NAME AS produto,
              pr.PROD_CATEGORY AS categoria,
              pr.PROD_CLASIF_2 AS litragem,
              pr.PROD_MANUFACTURER AS fabricante,
              pr.PROD_BRAND AS marca,
              ROUND(SUM(v.SALES_UNITS), 0) AS unidades_scanntech,
              ROUND(SUM(v.SALES_UNITS), 0) AS total_unidades,
              ROUND(SUM(v.SALES_UNITS / NULLIF(ap.QTDE_CX, 0)), 0) AS caixas_vendidas,
              ROUND(SUM(v.GROSS_SELLOUT), 2) AS receita_total,
              ROUND(SUM(v.GROSS_SELLOUT) / NULLIF(SUM(v.SALES_UNITS / NULLIF(ap.QTDE_CX, 0)), 0), 2) AS preco_medio_caixa
            FROM vta v
            JOIN prd pr ON v.PROD_ID = pr.PROD_ID
            LEFT JOIN {PORTFOLIO_TABLE} ap
              ON UPPER(TRIM(COALESCE(pr.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE(ap.PROD_CATEGORY, '')))
             AND UPPER(TRIM(COALESCE(pr.PROD_CLASIF_2, ''))) = UPPER(TRIM(COALESCE(ap.LITRAGEM, '')))
            WHERE pr.PROD_CATEGORY IS NOT NULL
            GROUP BY pr.PROD_ID, pr.PROD_NAME, pr.PROD_CATEGORY, pr.PROD_CLASIF_2, pr.PROD_MANUFACTURER, pr.PROD_BRAND
            """,
            """
            DROP VIEW IF EXISTS vendas_por_mes
            """,
            """
            DROP TABLE IF EXISTS vendas_por_mes
            """,
            """
            CREATE TABLE vendas_por_mes AS
            SELECT
              DATE_FORMAT(STR_TO_DATE(CONCAT(MONTH_ID, '01'), '%Y%m%d'), '%Y-%m') AS mes,
              ROUND(SUM(v.SALES_UNITS), 0) AS unidades_scanntech,
              ROUND(SUM(v.SALES_UNITS / NULLIF(ap.QTDE_CX, 0)), 0) AS caixas_vendidas,
              ROUND(SUM(v.GROSS_SELLOUT), 2) AS receita_total
            FROM vta v
            JOIN prd pr ON v.PROD_ID = pr.PROD_ID
            LEFT JOIN {PORTFOLIO_TABLE} ap
              ON UPPER(TRIM(COALESCE(pr.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE(ap.PROD_CATEGORY, '')))
             AND UPPER(TRIM(COALESCE(pr.PROD_CLASIF_2, ''))) = UPPER(TRIM(COALESCE(ap.LITRAGEM, '')))
            WHERE pr.PROD_CATEGORY IS NOT NULL
            GROUP BY MONTH_ID
            """,
            """
            DROP VIEW IF EXISTS market_share_fabricante
            """,
            """
            DROP TABLE IF EXISTS market_share_fabricante
            """,
            """
            CREATE TABLE market_share_fabricante AS
            SELECT
              p.PROD_MANUFACTURER AS fabricante,
              COUNT(DISTINCT p.PROD_ID) AS skus,
              COUNT(DISTINCT v.PDV_ID) AS pdvs_presentes,
              ROUND(SUM(v.SALES_UNITS), 0) AS total_unidades,
              ROUND(SUM(v.GROSS_SELLOUT), 2) AS total_receita,
              ROUND(
                SUM(v.GROSS_SELLOUT) / NULLIF((SELECT SUM(GROSS_SELLOUT) FROM vta), 0) * 100,
                2
              ) AS market_share_pct
            FROM vta v
            JOIN prd p ON v.PROD_ID = p.PROD_ID
            WHERE p.PROD_MANUFACTURER IS NOT NULL
            GROUP BY p.PROD_MANUFACTURER
            ORDER BY total_receita DESC
            """,
            """
            DROP VIEW IF EXISTS concorrentes_por_categoria
            """,
            """
            DROP TABLE IF EXISTS concorrentes_por_categoria
            """,
            """
            CREATE TABLE concorrentes_por_categoria AS
            WITH totais AS (
              SELECT
                p.PROD_CATEGORY AS categoria,
                ROUND(SUM(v.GROSS_SELLOUT), 2) AS faturamento_total_categoria
              FROM vta v
              JOIN prd p ON v.PROD_ID = p.PROD_ID
              LEFT JOIN {PORTFOLIO_TABLE} ap
                ON UPPER(TRIM(COALESCE(p.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE(ap.PROD_CATEGORY, '')))
               AND UPPER(TRIM(COALESCE(p.PROD_CLASIF_2, ''))) = UPPER(TRIM(COALESCE(ap.LITRAGEM, '')))
              WHERE p.PROD_CATEGORY IS NOT NULL
              GROUP BY p.PROD_CATEGORY
            ),
            base AS (
              SELECT
                p.PROD_CATEGORY AS categoria,
                COALESCE(NULLIF(p.PROD_MANUFACTURER, ''), NULLIF(p.PROD_BRAND, ''), 'SEM FABRICANTE') AS concorrente,
                ROUND(SUM(v.SALES_UNITS), 0) AS unidades_scanntech,
                ROUND(SUM(v.SALES_UNITS / NULLIF(ap.QTDE_CX, 0)), 0) AS caixas_vendidas,
                ROUND(SUM(v.GROSS_SELLOUT), 2) AS faturamento
              FROM vta v
              JOIN prd p ON v.PROD_ID = p.PROD_ID
              LEFT JOIN {PORTFOLIO_TABLE} ap
                ON UPPER(TRIM(COALESCE(p.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE(ap.PROD_CATEGORY, '')))
               AND UPPER(TRIM(COALESCE(p.PROD_CLASIF_2, ''))) = UPPER(TRIM(COALESCE(ap.LITRAGEM, '')))
              WHERE p.PROD_CATEGORY IS NOT NULL
                AND UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) <> 'AQUAFAST'
              GROUP BY p.PROD_CATEGORY, COALESCE(NULLIF(p.PROD_MANUFACTURER, ''), NULLIF(p.PROD_BRAND, ''), 'SEM FABRICANTE')
            )
            SELECT
              base.categoria,
              base.concorrente,
              base.faturamento,
              base.unidades_scanntech,
              base.caixas_vendidas,
              ROUND(base.faturamento / NULLIF(totais.faturamento_total_categoria, 0) * 100, 2) AS participacao_categoria_pct,
              ROW_NUMBER() OVER (
                PARTITION BY base.categoria
                ORDER BY base.faturamento DESC, base.unidades_scanntech DESC, base.concorrente
              ) AS ranking_categoria
            FROM base
            JOIN totais ON totais.categoria = base.categoria
            ORDER BY base.categoria, ranking_categoria, base.concorrente
            """,
            """
            DROP VIEW IF EXISTS share_aquafast_por_categoria
            """,
            """
            DROP TABLE IF EXISTS share_aquafast_por_categoria
            """,
            """
            CREATE TABLE share_aquafast_por_categoria AS
            SELECT
              p.PROD_CATEGORY AS categoria,
              ROUND(SUM(v.GROSS_SELLOUT), 2) AS faturamento_total_categoria,
              ROUND(SUM(CASE WHEN UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) = 'AQUAFAST' THEN v.GROSS_SELLOUT ELSE 0 END), 2) AS faturamento_aquafast,
              ROUND(SUM(CASE WHEN UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) <> 'AQUAFAST' THEN v.GROSS_SELLOUT ELSE 0 END), 2) AS faturamento_concorrentes,
              ROUND(
                SUM(CASE WHEN UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) = 'AQUAFAST' THEN v.GROSS_SELLOUT ELSE 0 END)
                / NULLIF(SUM(v.GROSS_SELLOUT), 0) * 100,
                2
              ) AS share_aquafast_pct,
              ROUND(SUM(CASE WHEN UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) = 'AQUAFAST' THEN v.SALES_UNITS ELSE 0 END), 0) AS unidades_aquafast,
              ROUND(SUM(CASE WHEN UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) <> 'AQUAFAST' THEN v.SALES_UNITS ELSE 0 END), 0) AS unidades_concorrentes
            FROM vta v
            JOIN prd p ON v.PROD_ID = p.PROD_ID
            WHERE p.PROD_CATEGORY IS NOT NULL
            GROUP BY p.PROD_CATEGORY
            ORDER BY share_aquafast_pct DESC, faturamento_total_categoria DESC, categoria
            """,
            """
            DROP VIEW IF EXISTS top_concorrentes_por_cidade
            """,
            """
            DROP TABLE IF EXISTS top_concorrentes_por_cidade
            """,
            """
            CREATE TABLE top_concorrentes_por_cidade AS
            WITH base AS (
              SELECT
                COALESCE(d.PDV_LOCATION, 'SEM CIDADE') AS cidade,
                COALESCE(d.PDV_STATE, 'SEM UF') AS uf,
                COALESCE(NULLIF(p.PROD_MANUFACTURER, ''), NULLIF(p.PROD_BRAND, ''), 'SEM FABRICANTE') AS concorrente,
                ROUND(SUM(v.GROSS_SELLOUT), 2) AS faturamento,
                ROUND(SUM(v.SALES_UNITS), 0) AS unidades_scanntech
              FROM vta v
              JOIN pdv d ON v.PDV_ID = d.PDV_ID
              JOIN prd p ON v.PROD_ID = p.PROD_ID
              WHERE p.PROD_CATEGORY IS NOT NULL
                AND UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) <> 'AQUAFAST'
                AND d.PDV_LOCATION IS NOT NULL
              GROUP BY
                COALESCE(d.PDV_LOCATION, 'SEM CIDADE'),
                COALESCE(d.PDV_STATE, 'SEM UF'),
                COALESCE(NULLIF(p.PROD_MANUFACTURER, ''), NULLIF(p.PROD_BRAND, ''), 'SEM FABRICANTE')
            ),
            ranked AS (
              SELECT
                cidade,
                uf,
                concorrente,
                faturamento,
                unidades_scanntech,
                ROW_NUMBER() OVER (
                  PARTITION BY cidade, uf
                  ORDER BY faturamento DESC, unidades_scanntech DESC, concorrente
                ) AS ranking_cidade
              FROM base
            )
            SELECT
              cidade,
              uf,
              concorrente,
              faturamento,
              unidades_scanntech,
              ranking_cidade
            FROM ranked
            ORDER BY cidade, ranking_cidade, concorrente
            """,
            """
            DROP VIEW IF EXISTS concorrentes_crescimento_90_dias
            """,
            """
            DROP TABLE IF EXISTS concorrentes_crescimento_90_dias
            """,
            """
            CREATE TABLE concorrentes_crescimento_90_dias AS
            WITH meses AS (
              SELECT DISTINCT MONTH_ID
              FROM vta
              WHERE MONTH_ID IS NOT NULL
              ORDER BY MONTH_ID DESC
              LIMIT 6
            ),
            base AS (
              SELECT
                COALESCE(NULLIF(p.PROD_MANUFACTURER, ''), NULLIF(p.PROD_BRAND, ''), 'SEM FABRICANTE') AS concorrente,
                p.PROD_CATEGORY AS categoria,
                v.SALES_UNITS AS unidades,
                v.GROSS_SELLOUT AS receita,
                DENSE_RANK() OVER (ORDER BY v.MONTH_ID DESC) AS month_rank
              FROM vta v
              JOIN prd p ON v.PROD_ID = p.PROD_ID
              JOIN meses m ON m.MONTH_ID = v.MONTH_ID
              WHERE p.PROD_CATEGORY IS NOT NULL
                AND UPPER(TRIM(COALESCE(p.PROD_MANUFACTURER, ''))) <> 'AQUAFAST'
            ),
            agg AS (
              SELECT
                concorrente,
                categoria,
                ROUND(SUM(CASE WHEN month_rank <= 3 THEN receita ELSE 0 END), 2) AS faturamento_90d,
                ROUND(SUM(CASE WHEN month_rank BETWEEN 4 AND 6 THEN receita ELSE 0 END), 2) AS faturamento_90d_anterior,
                ROUND(SUM(CASE WHEN month_rank <= 3 THEN unidades ELSE 0 END), 0) AS unidades_90d,
                ROUND(SUM(CASE WHEN month_rank BETWEEN 4 AND 6 THEN unidades ELSE 0 END), 0) AS unidades_90d_anterior
              FROM base
              GROUP BY concorrente, categoria
            )
            SELECT
              concorrente,
              categoria,
              faturamento_90d,
              faturamento_90d_anterior,
              ROUND(faturamento_90d - faturamento_90d_anterior, 2) AS variacao_abs,
              ROUND((faturamento_90d - faturamento_90d_anterior) / NULLIF(faturamento_90d_anterior, 0) * 100, 2) AS variacao_pct,
              unidades_90d,
              unidades_90d_anterior
            FROM agg
            ORDER BY variacao_abs DESC, faturamento_90d DESC, concorrente, categoria
            """,
            """
            DROP VIEW IF EXISTS vendas_por_estado
            """,
            """
            DROP TABLE IF EXISTS vendas_por_estado
            """,
            """
            CREATE TABLE vendas_por_estado AS
            SELECT
              d.PDV_STATE AS estado,
              COUNT(DISTINCT d.PDV_ID) AS total_pdvs,
              COUNT(DISTINCT p.PROD_MANUFACTURER) AS fabricantes,
              ROUND(SUM(v.SALES_UNITS), 0) AS unidades_scanntech,
              ROUND(SUM(v.SALES_UNITS / NULLIF(ap.QTDE_CX, 0)), 0) AS caixas_vendidas,
              ROUND(SUM(v.GROSS_SELLOUT), 2) AS receita_total
            FROM vta v
            JOIN pdv d ON v.PDV_ID = d.PDV_ID
            JOIN prd p ON v.PROD_ID = p.PROD_ID
            LEFT JOIN {PORTFOLIO_TABLE} ap
              ON UPPER(TRIM(COALESCE(p.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE(ap.PROD_CATEGORY, '')))
             AND UPPER(TRIM(COALESCE(p.PROD_CLASIF_2, ''))) = UPPER(TRIM(COALESCE(ap.LITRAGEM, '')))
            WHERE d.PDV_STATE IS NOT NULL
            GROUP BY d.PDV_STATE
            ORDER BY receita_total DESC
            """,
            """
            DROP VIEW IF EXISTS vendas_por_cidade
            """,
            """
            DROP TABLE IF EXISTS vendas_por_cidade
            """,
            """
            CREATE TABLE vendas_por_cidade AS
            SELECT
              COALESCE(d.PDV_LOCATION, 'SEM CIDADE') AS cidade,
              COALESCE(d.PDV_STATE, 'SEM UF') AS estado,
              COUNT(DISTINCT d.PDV_ID) AS pdvs,
              ROUND(SUM(v.SALES_UNITS), 0) AS unidades_scanntech,
              ROUND(SUM(v.SALES_UNITS / NULLIF(ap.QTDE_CX, 0)), 0) AS caixas_vendidas,
              ROUND(SUM(v.GROSS_SELLOUT), 2) AS receita_total
            FROM vta v
            JOIN pdv d ON v.PDV_ID = d.PDV_ID
            JOIN prd p ON v.PROD_ID = p.PROD_ID
            LEFT JOIN {PORTFOLIO_TABLE} ap
              ON UPPER(TRIM(COALESCE(p.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE(ap.PROD_CATEGORY, '')))
             AND UPPER(TRIM(COALESCE(p.PROD_CLASIF_2, ''))) = UPPER(TRIM(COALESCE(ap.LITRAGEM, '')))
            WHERE d.PDV_LOCATION IS NOT NULL
              AND p.PROD_CATEGORY IS NOT NULL
            GROUP BY COALESCE(d.PDV_LOCATION, 'SEM CIDADE'), COALESCE(d.PDV_STATE, 'SEM UF')
            ORDER BY receita_total DESC, caixas_vendidas DESC, cidade
            """,
            """
            DROP VIEW IF EXISTS ranking_redes
            """,
            """
            DROP TABLE IF EXISTS ranking_redes
            """,
            """
            CREATE TABLE ranking_redes AS
            SELECT
              COALESCE(d.PDV_STORE_CHAIN, 'SEM REDE') AS rede,
              d.STORE_CLASSIFICATION AS tipo_loja,
              COUNT(DISTINCT d.PDV_ID) AS total_lojas,
              COUNT(DISTINCT p.PROD_MANUFACTURER) AS fabricantes,
              ROUND(SUM(v.SALES_UNITS), 0) AS total_unidades,
              ROUND(SUM(v.GROSS_SELLOUT), 2) AS total_receita
            FROM vta v
            JOIN pdv d ON v.PDV_ID = d.PDV_ID
            JOIN prd p ON v.PROD_ID = p.PROD_ID
            GROUP BY COALESCE(d.PDV_STORE_CHAIN, 'SEM REDE'), d.STORE_CLASSIFICATION
            ORDER BY total_receita DESC
            """,
            """
            DROP VIEW IF EXISTS top_produtos_categoria
            """,
            """
            DROP TABLE IF EXISTS top_produtos_categoria
            """,
            """
            CREATE TABLE top_produtos_categoria AS
            SELECT
              p.PROD_CATEGORY AS categoria,
              p.PROD_NAME AS produto,
              p.PROD_MANUFACTURER AS fabricante,
              p.PROD_BRAND AS marca,
              p.PROD_BARCODE AS ean,
              COUNT(DISTINCT v.PDV_ID) AS pdvs_com_venda,
              ROUND(SUM(v.SALES_UNITS), 0) AS unidades_scanntech,
              ROUND(SUM(v.SALES_UNITS), 0) AS total_unidades,
              ROUND(SUM(v.SALES_UNITS / NULLIF(ap.QTDE_CX, 0)), 0) AS caixas_vendidas,
              ROUND(SUM(v.GROSS_SELLOUT), 2) AS total_receita,
              ROUND(SUM(v.GROSS_SELLOUT) / NULLIF(SUM(v.SALES_UNITS / NULLIF(ap.QTDE_CX, 0)), 0), 2) AS preco_medio_caixa
            FROM vta v
            JOIN prd p ON v.PROD_ID = p.PROD_ID
            LEFT JOIN {PORTFOLIO_TABLE} ap
              ON UPPER(TRIM(COALESCE(p.PROD_CATEGORY, ''))) = UPPER(TRIM(COALESCE(ap.PROD_CATEGORY, '')))
             AND UPPER(TRIM(COALESCE(p.PROD_CLASIF_2, ''))) = UPPER(TRIM(COALESCE(ap.LITRAGEM, '')))
            WHERE p.PROD_CATEGORY IS NOT NULL
            GROUP BY
              p.PROD_CATEGORY,
              p.PROD_NAME,
              p.PROD_MANUFACTURER,
              p.PROD_BRAND,
              p.PROD_BARCODE
            ORDER BY total_receita DESC
            """,
        ]
        for statement in statements:
            cur.execute(statement)
        _create_product_resolution_table(con)
        con.commit()
        cur.close()
        _SCHEMA_BOOTSTRAPPED = True
    finally:
        con.close()


def open_connection():
    if CHAT_BACKEND == "mysql":
        _bootstrap_compatibility_views()
        return mysql.connector.connect(**_mysql_config())
    if not DUCKDB_PATH.exists():
        raise FileNotFoundError(f"Banco nao encontrado: {DUCKDB_PATH}")
    return duckdb.connect(str(DUCKDB_PATH), read_only=True)


def ensure_read_only_sql(sql: str) -> str:
    candidate = sql.strip().rstrip(";")
    if not re.match(r"^(select|with|show|describe)\b", candidate, flags=re.IGNORECASE):
        raise ValueError("A API aceita apenas consultas de leitura.")
    return candidate


def _sql_supports_row_cap(sql: str) -> bool:
    head = sql.strip().lstrip("(").lower()
    return head.startswith("select") or head.startswith("with")


def _strip_trailing_order_by(sql: str) -> str:
    candidate = sql.strip().rstrip(";")
    match = re.search(r"\border\s+by\b", candidate, flags=re.IGNORECASE)
    if not match:
        return candidate
    return candidate[: match.start()].rstrip()


def _execute_sql(con: Any, sql: str) -> tuple[list[str], list[tuple[Any, ...]]]:
    if CHAT_BACKEND == "mysql":
        cur = con.cursor()
        try:
            cur.execute(sql)
            columns = [item[0] for item in cur.description] if cur.description else []
            rows = cur.fetchall()
            normalized_rows = [tuple(_normalize_value(value) for value in row) for row in rows]
            return columns, normalized_rows
        finally:
            cur.close()

    result = con.execute(sql)
    columns = [item[0] for item in result.description]
    rows = result.fetchall()
    normalized_rows = [tuple(_normalize_value(value) for value in row) for row in rows]
    return columns, normalized_rows


def run_query(sql: str, *, row_cap: int | None = None) -> dict[str, Any]:
    candidate = ensure_read_only_sql(sql.strip().rstrip(";"))
    if _is_audit_products_without_subgroup_query(candidate):
        con = open_connection()
        try:
            result = _audit_build_result(con)
        finally:
            con.close()
        if row_cap is not None and row_cap > 0 and len(result["rows"]) > row_cap:
            result = {**result, "rows": result["rows"][:row_cap], "truncated": True, "row_cap": row_cap}
            result["row_count"] = len(result["rows"])
            result["markdown"] = format_markdown(result["columns"], result["rows"])
        return result

    if _is_lojas_com_concorrente_sem_aquafast_query(candidate):
        result = _build_lojas_com_concorrente_sem_aquafast_result()
        if row_cap is not None and row_cap > 0 and len(result["rows"]) > row_cap:
            result = {**result, "rows": result["rows"][:row_cap], "truncated": True, "row_cap": row_cap}
            result["row_count"] = len(result["rows"])
            result["markdown"] = format_markdown(result["columns"], result["rows"])
        return result

    con = open_connection()
    try:
        truncated = False
        if row_cap is not None and row_cap > 0 and _sql_supports_row_cap(candidate):
            fetch_limit = int(row_cap) + 1
            wrapped = f"SELECT * FROM ({candidate}) AS _aquafast_sub LIMIT {fetch_limit}"
            columns, rows = _execute_sql(con, wrapped)
        else:
            columns, rows = _execute_sql(con, candidate)
        if row_cap is not None and row_cap > 0 and len(rows) > row_cap:
            truncated = True
            rows = rows[:row_cap]

        return {
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
            "markdown": format_markdown(columns, rows),
            "truncated": truncated,
            "row_cap": row_cap,
        }
    finally:
        con.close()


def clamp_page(value: int) -> int:
    return max(1, int(value))


def clamp_page_size(value: int) -> int:
    return max(1, min(int(value), REPORT_PAGE_SIZE_LIMIT))


def get_report_spec(report_name: str) -> dict[str, Any]:
    try:
        return REPORT_SPECS[report_name]
    except KeyError as exc:
        raise KeyError(f"Relatorio desconhecido: {report_name}") from exc


def list_report_specs() -> list[dict[str, Any]]:
    return [
        {
            "name": name,
            "title": spec["title"],
            "description": spec["description"],
            "default_page_size": 50,
        }
        for name, spec in REPORT_SPECS.items()
    ]


def run_report(report_name: str, page: int = 1, page_size: int = 50) -> dict[str, Any]:
    if report_name == "historico_consultas":
        spec = get_report_spec(report_name)
        page = clamp_page(page)
        page_size = clamp_page_size(page_size)
        full_result = _build_historico_consultas_result()
        total_rows = full_result["row_count"]
        total_pages = math.ceil(total_rows / page_size) if total_rows else 0
        offset = (page - 1) * page_size
        page_rows = full_result["rows"][offset : offset + page_size]
        page_result = {
            **full_result,
            "rows": page_rows,
            "row_count": len(page_rows),
            "markdown": format_markdown(full_result["columns"], page_rows),
            "truncated": total_rows > len(page_rows),
            "row_cap": page_size,
        }
        page_result = _attach_history_metadata(
            page_result,
            pergunta=spec["title"],
            report_name=report_name,
            rows_returned=len(page_rows),
        )
        return {
            "report_name": report_name,
            "title": spec["title"],
            "description": spec["description"],
            "sql": spec["sql"],
            "base_sql": spec["sql"],
            "page": page,
            "page_size": page_size,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "has_previous_page": page > 1,
            "has_next_page": page < total_pages,
            **page_result,
        }

    if report_name == "lojas_com_concorrente_sem_aquafast":
        spec = get_report_spec(report_name)
        page = clamp_page(page)
        page_size = clamp_page_size(page_size)
        full_result = _build_lojas_com_concorrente_sem_aquafast_result()
        total_rows = full_result["row_count"]
        total_pages = math.ceil(total_rows / page_size) if total_rows else 0
        offset = (page - 1) * page_size
        page_rows = full_result["rows"][offset : offset + page_size]
        page_result = {
            **full_result,
            "rows": page_rows,
            "row_count": len(page_rows),
            "markdown": format_markdown(full_result["columns"], page_rows),
            "truncated": total_rows > len(page_rows),
            "row_cap": page_size,
        }
        page_result = _attach_history_metadata(
            page_result,
            pergunta=spec["title"],
            report_name=report_name,
            rows_returned=len(page_rows),
        )
        return {
            "report_name": report_name,
            "title": spec["title"],
            "description": spec["description"],
            "sql": spec["sql"],
            "base_sql": spec["sql"],
            "page": page,
            "page_size": page_size,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "has_previous_page": page > 1,
            "has_next_page": page < total_pages,
            **page_result,
        }

    spec = get_report_spec(report_name)
    page = clamp_page(page)
    page_size = clamp_page_size(page_size)
    base_sql = ensure_read_only_sql(spec["sql"])

    if report_name == AUDIT_PRODUCTS_SENTINEL:
        con = open_connection()
        try:
            full_result = _audit_build_result(con)
        finally:
            con.close()
        total_rows = full_result["row_count"]
        total_pages = math.ceil(total_rows / page_size) if total_rows else 0
        offset = (page - 1) * page_size
        page_rows = full_result["rows"][offset : offset + page_size]
        page_result = {
            **full_result,
            "rows": page_rows,
            "row_count": len(page_rows),
            "markdown": format_markdown(full_result["columns"], page_rows),
            "truncated": total_rows > len(page_rows),
            "row_cap": page_size,
        }
        page_result = _attach_history_metadata(
            page_result,
            pergunta=spec["title"],
            report_name=report_name,
            rows_returned=len(page_rows),
        )
        return {
            "report_name": report_name,
            "title": spec["title"],
            "description": spec["description"],
            "sql": base_sql,
            "base_sql": base_sql,
            "page": page,
            "page_size": page_size,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "has_previous_page": page > 1,
            "has_next_page": page < total_pages,
            **page_result,
        }

    offset = (page - 1) * page_size
    count_sql = f"SELECT COUNT(*) AS total_rows FROM ({_strip_trailing_order_by(base_sql)}) AS report_data"
    page_sql = f"SELECT * FROM ({base_sql}) AS report_data LIMIT {page_size} OFFSET {offset}"

    total_rows_result = run_query(count_sql)
    total_rows = int(total_rows_result["rows"][0][0]) if total_rows_result["rows"] else 0
    total_pages = math.ceil(total_rows / page_size) if total_rows else 0

    page_result = run_query(page_sql)
    page_result = _attach_history_metadata(
        page_result,
        pergunta=spec["title"],
        report_name=report_name,
        rows_returned=int(page_result.get("row_count", 0)),
    )
    return {
        "report_name": report_name,
        "title": spec["title"],
        "description": spec["description"],
        "sql": page_sql,
        "base_sql": base_sql,
        "page": page,
        "page_size": page_size,
        "total_rows": total_rows,
        "total_pages": total_pages,
        "has_previous_page": page > 1,
        "has_next_page": page < total_pages,
        **page_result,
    }


def write_xlsx_report(title: str, columns: list[str], rows: list[tuple[Any, ...]]) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{sanitize_filename(title)}_{stamp}.xlsx"
    file_path = EXPORT_DIR / filename

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resultado"
    sheet.append(columns)
    for row in rows:
        sheet.append(list(row))

    for idx, column in enumerate(columns, start=1):
        width = max(len(str(column)), 14)
        sheet.column_dimensions[get_column_letter(idx)].width = min(width + 2, 42)

    workbook.save(file_path)
    return file_path


def export_query(sql: str, title: str) -> dict[str, Any]:
    sql = ensure_read_only_sql(sql)
    result = run_query(sql, row_cap=EXPORT_RESULT_ROW_CAP)
    file_path = write_xlsx_report(title, result["columns"], result["rows"])
    return {
        **result,
        "file_name": file_path.name,
        "file_path": str(file_path),
        "download_url": f"http://localhost:8001/download/{file_path.name}",
    }


def legacy_question_to_sql(question: str) -> tuple[str, str]:
    route = resolve_official_route(question)
    if route is not None:
        return route.title, route.sql

    q = normalize_business_question(question)

    if any(
        term in q
        for term in [
            "produtos sem subgrupo cigam",
            "produtos sem padronizacao",
            "auditoria produtos sem padronizacao",
            "auditoria produtos sem padronização",
            "quais produtos nao casam com o portfolio",
            "quais produtos nao casam com o portifolio",
        ]
    ):
        return (
            "Auditoria produtos sem SUBGRUPO_CIGAM",
            """
            SELECT *
            FROM auditoria_produtos_sem_subgrupo_cigam
            ORDER BY faturamento DESC, caixas_vendidas DESC, ocorrencias DESC, produto_original_scanntech
            """.strip(),
        )

    if any(term in q for term in ["historico de consultas", "historico consultas", "ultimas consultas", "quais relatorios eu consultei"]):
        return (
            "Historico de consultas",
            """
            SELECT
                timestamp AS data_hora,
                pergunta,
                report_name AS relatorio,
                metric,
                rows_returned AS linhas_retornadas,
                status
            FROM aquafast_query_history
            ORDER BY timestamp DESC, id DESC
            LIMIT 20
            """.strip(),
        )

    if any(term in q for term in ["ticket medio por cliente", "ticket por cliente", "ticket medio por cliente da aquafast"]):
        return (
            "Ticket medio por cliente Aquafast",
            """
            SELECT cliente, ticket_medio_caixa, unidades_scanntech, caixas_vendidas, receita_total, primeira_compra, ultima_compra
            FROM ranking_clientes
            ORDER BY ticket_medio_caixa DESC, receita_total DESC, cliente
            LIMIT 20
            """.strip(),
        )

    if any(term in q for term in ["ticket medio por produto", "ticket por produto", "ticket medio por produto da aquafast"]):
        return (
            "Ticket medio por produto Aquafast",
            """
            SELECT
                produto_padrao,
                produto_original_exemplo,
                subgrupo_cigam,
                variacoes_produto_original,
                preco_medio_caixa AS ticket_medio_caixa,
                unidades_scanntech,
                caixas_vendidas,
                receita_total,
                categoria,
                fabricante,
                marca
            FROM ranking_produtos
            ORDER BY preco_medio_caixa DESC, receita_total DESC, produto_padrao
            LIMIT 20
            """.strip(),
        )

    if any(term in q for term in ["ultima venda por produto", "ultimo valor por produto", "ultimo pedido por produto", "ultima venda do produto"]):
        return (
            "Ultima venda por produto Aquafast",
            """
            SELECT
                codigo,
                produto,
                cliente,
                data_venda,
                qtd,
                valor_unitario,
                valor_total
            FROM (
                SELECT
                    COD_PRODUTO AS codigo,
                    DESC_PRODUTO AS produto,
                    RAZAO_SOCIAL AS cliente,
                    DATA_VENDA AS data_venda,
                    ROUND(QTD, 0) AS qtd,
                    ROUND(VALOR_UNITARIO, 2) AS valor_unitario,
                    ROUND(VALOR_TOTAL, 2) AS valor_total,
                    ROW_NUMBER() OVER (PARTITION BY COD_PRODUTO ORDER BY DATA_VENDA DESC, COD_PRODUTO DESC, CNPJ DESC) AS rn
                FROM scanntech
            ) last_sale
            WHERE rn = 1
            ORDER BY data_venda DESC, produto
            LIMIT 50
            """.strip(),
        )

    if any(term in q for term in ["primeira e ultima venda da base", "primeira venda da base", "ultima venda da base", "primeira e ultima venda"]):
        return (
            "Primeira e ultima venda da base Aquafast",
            """
            WITH ranked AS (
                SELECT
                    COD_PRODUTO AS codigo,
                    DESC_PRODUTO AS produto,
                    RAZAO_SOCIAL AS cliente,
                    DATA_VENDA AS data_venda,
                    ROUND(QTD, 0) AS qtd,
                    ROUND(VALOR_UNITARIO, 2) AS valor_unitario,
                    ROUND(VALOR_TOTAL, 2) AS valor_total,
                    ROW_NUMBER() OVER (ORDER BY DATA_VENDA ASC, COD_PRODUTO ASC, CNPJ ASC) AS rn_first,
                    ROW_NUMBER() OVER (ORDER BY DATA_VENDA DESC, COD_PRODUTO DESC, CNPJ DESC) AS rn_last
                FROM scanntech
            )
            SELECT
                'primeira venda' AS marco,
                codigo,
                produto,
                cliente,
                data_venda,
                qtd,
                valor_unitario,
                valor_total
            FROM ranked
            WHERE rn_first = 1
            UNION ALL
            SELECT
                'ultima venda' AS marco,
                codigo,
                produto,
                cliente,
                data_venda,
                qtd,
                valor_unitario,
                valor_total
            FROM ranked
            WHERE rn_last = 1
            ORDER BY data_venda
            """.strip(),
        )

    if any(term in q for term in ["curva abc de produtos", "abc de produtos", "abc produtos"]):
        return (
            "Curva ABC de produtos Aquafast",
            """
            WITH base AS (
                SELECT
                    produto_padrao,
                    produto_original_exemplo,
                    subgrupo_cigam,
                    variacoes_produto_original,
                    categoria,
                    fabricante,
                    marca,
                    unidades_scanntech,
                    caixas_vendidas,
                    receita_total,
                    SUM(receita_total) OVER () AS receita_geral,
                    SUM(receita_total) OVER (
                        ORDER BY receita_total DESC, produto_padrao
                        ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                    ) AS receita_acumulada
                FROM ranking_produtos
            )
            SELECT
                produto_padrao,
                produto_original_exemplo,
                subgrupo_cigam,
                variacoes_produto_original,
                categoria,
                fabricante,
                marca,
                unidades_scanntech,
                caixas_vendidas,
                receita_total,
                ROUND(receita_total / NULLIF(receita_geral, 0) * 100, 2) AS percentual_receita,
                ROUND(receita_acumulada / NULLIF(receita_geral, 0) * 100, 2) AS percentual_acumulado,
                CASE
                    WHEN receita_acumulada / NULLIF(receita_geral, 0) <= 0.80 THEN 'A'
                    WHEN receita_acumulada / NULLIF(receita_geral, 0) <= 0.95 THEN 'B'
                    ELSE 'C'
                END AS classe_abc
            FROM base
            ORDER BY receita_total DESC, produto_padrao
            LIMIT 50
            """.strip(),
        )

    if any(term in q for term in ["curva abc de clientes", "abc de clientes", "abc clientes"]):
        return (
            "Curva ABC de clientes Aquafast",
            """
            WITH base AS (
                SELECT
                    cliente,
                    unidades_scanntech,
                    caixas_vendidas,
                    receita_total,
                    ticket_medio_caixa,
                    primeira_compra,
                    ultima_compra,
                    SUM(receita_total) OVER () AS receita_geral,
                    SUM(receita_total) OVER (
                        ORDER BY receita_total DESC, cliente
                        ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                    ) AS receita_acumulada
                FROM ranking_clientes
            )
            SELECT
                cliente,
                unidades_scanntech,
                caixas_vendidas,
                receita_total,
                ticket_medio_caixa,
                primeira_compra,
                ultima_compra,
                ROUND(receita_total / NULLIF(receita_geral, 0) * 100, 2) AS percentual_receita,
                ROUND(receita_acumulada / NULLIF(receita_geral, 0) * 100, 2) AS percentual_acumulado,
                CASE
                    WHEN receita_acumulada / NULLIF(receita_geral, 0) <= 0.80 THEN 'A'
                    WHEN receita_acumulada / NULLIF(receita_geral, 0) <= 0.95 THEN 'B'
                    ELSE 'C'
                END AS classe_abc
            FROM base
            ORDER BY receita_total DESC, cliente
            LIMIT 50
            """.strip(),
        )

    m_top_cli = re.search(r"\btop\s+(\d+)\s+clientes", q)
    if m_top_cli:
        n = min(max(int(m_top_cli.group(1)), 1), 200)
        return (
            f"Top {n} clientes Aquafast por caixa",
            f"SELECT * FROM ranking_clientes ORDER BY caixas_vendidas DESC, receita_total DESC, cliente LIMIT {n}",
        )

    m_top_prd = re.search(r"\btop\s+(\d+)\s+produtos", q)
    if m_top_prd:
        n = min(max(int(m_top_prd.group(1)), 1), 200)
        return (
            f"Top {n} produtos Aquafast por caixa",
            f"SELECT * FROM ranking_produtos ORDER BY total_vendas DESC, receita_total DESC, produto LIMIT {n}",
        )

    if any(term in q for term in ["concorrentes por categoria", "concorrente por categoria", "qual concorrente domina cada categoria", "qual concorrente domina categoria"]):
        return (
            "Concorrentes por categoria",
            """
            SELECT *
            FROM concorrentes_por_categoria
            ORDER BY categoria, ranking_categoria
            """.strip(),
        )

    if any(term in q for term in ["share aquafast por categoria", "participacao aquafast por categoria", "participacao da aquafast por categoria", "qual a participacao da aquafast por categoria"]):
        return (
            "Share Aquafast por categoria",
            """
            SELECT *
            FROM share_aquafast_por_categoria
            ORDER BY share_aquafast_pct DESC, faturamento_total_categoria DESC, categoria
            """.strip(),
        )

    if any(term in q for term in ["lojas com concorrente sem aquafast", "onde concorrente vende e aquafast nao", "quais lojas vendem concorrente mas nao vendem aquafast"]):
        return (
            "Lojas com concorrente sem Aquafast",
            """
            SELECT *
            FROM lojas_com_concorrente_sem_aquafast
            ORDER BY faturamento_concorrente DESC, unidades_scanntech DESC, loja, categoria, concorrente
            """.strip(),
        )

    if any(term in q for term in ["top concorrentes por cidade", "concorrentes por cidade"]):
        return (
            "Top concorrentes por cidade",
            """
            SELECT *
            FROM top_concorrentes_por_cidade
            ORDER BY cidade, ranking_cidade
            """.strip(),
        )

    if any(term in q for term in ["concorrentes crescimento 90 dias", "concorrentes em crescimento 90 dias", "qual concorrente mais cresce nos ultimos 90 dias"]):
        return (
            "Concorrentes em crescimento 90 dias",
            """
            SELECT *
            FROM concorrentes_crescimento_90_dias
            ORDER BY variacao_pct DESC, faturamento_90d DESC, concorrente, categoria
            """.strip(),
        )

    if any(term in q for term in ["maior concorrente", "maiores concorrentes", "concorrente", "concorrentes", "concorrencia", "competidor", "competidores"]):
        limit = 1 if any(term in q for term in ["maior concorrente", "principal concorrente"]) else 10
        return (
            "Maiores concorrentes da Aquafast",
            f"""
            SELECT
              fabricante,
              skus,
              pdvs,
              total_unidades,
              total_receita,
              market_share_pct
            FROM ms_mercado_aquafast
            WHERE LOWER(fabricante) <> 'aquafast'
            ORDER BY total_receita DESC, market_share_pct DESC, fabricante
            LIMIT {limit}
            """.strip(),
        )

    if any(term in q for term in ["top 20 clientes", "clientes por valor", "ranking clientes", "clientes que mais compraram", "valor total dos clientes"]):
        return "Top clientes Aquafast por caixa", "SELECT * FROM ranking_clientes ORDER BY caixas_vendidas DESC, receita_total DESC, cliente LIMIT 20"

    if any(term in q for term in ["clientes com maior faturamento", "faturamento por cliente", "ranking de clientes por faturamento", "clientes aquafast com maior faturamento"]):
        return (
            "Clientes com maior faturamento Aquafast",
            """
            SELECT cliente, unidades_scanntech, caixas_vendidas, receita_total, ticket_medio_caixa, primeira_compra, ultima_compra
            FROM ranking_clientes
            ORDER BY receita_total DESC, caixas_vendidas DESC, cliente
            LIMIT 20
            """.strip(),
        )

    if any(term in q for term in ["clientes com menor compra", "menor compra", "clientes aquafast com menor compra"]):
        return (
            "Clientes com menor compra Aquafast",
            """
            SELECT cliente, unidades_scanntech, caixas_vendidas, receita_total, ticket_medio_caixa, primeira_compra, ultima_compra
            FROM ranking_clientes
            ORDER BY caixas_vendidas ASC, receita_total ASC, cliente
            LIMIT 20
            """.strip(),
        )

    if any(term in q for term in ["ticket medio por cliente", "ticket por cliente", "ticket medio por cliente da aquafast"]):
        return (
            "Ticket medio por cliente Aquafast",
            """
            SELECT cliente, ticket_medio_caixa, unidades_scanntech, caixas_vendidas, receita_total, primeira_compra, ultima_compra
            FROM ranking_clientes
            ORDER BY ticket_medio_caixa DESC, receita_total DESC, cliente
            LIMIT 20
            """.strip(),
        )

    if any(term in q for term in ["top 20 produtos", "produtos mais vendidos", "ranking produtos"]):
        return "Top produtos Aquafast por caixa", "SELECT * FROM ranking_produtos ORDER BY total_vendas DESC, receita_total DESC, produto LIMIT 20"

    if any(term in q for term in ["produtos por faturamento", "produtos aquafast por faturamento", "faturamento por produto", "produtos por receita", "ranking de receita por produto", "ranking de faturamento"]):
        return (
            "Receita por produto Aquafast (top 30)",
            """
            SELECT
                produto_padrao,
                produto_original_exemplo,
                subgrupo_cigam,
                variacoes_produto_original,
                unidades_scanntech,
                caixas_vendidas,
                receita_total,
                categoria,
                fabricante,
                marca
            FROM ranking_produtos
            ORDER BY receita_total DESC, caixas_vendidas DESC, produto_padrao
            LIMIT 30
            """.strip(),
        )

    if any(
        term in q
        for term in [
            "produtos por quantidade",
            "produtos aquafast por quantidade",
            "mais vendidos em quantidade",
            "maior volume de vendas",
            "produtos com maior volume",
            "mais unidades vendidas",
        ]
    ):
        return (
            "Produtos com maior volume (caixas)",
            """
            SELECT
                produto_padrao,
                produto_original_exemplo,
                subgrupo_cigam,
                variacoes_produto_original,
                unidades_scanntech,
                caixas_vendidas,
                receita_total,
                categoria,
                fabricante,
                marca
            FROM ranking_produtos
            ORDER BY unidades_scanntech DESC, caixas_vendidas DESC, receita_total DESC, produto_padrao
            LIMIT 50
            """.strip(),
        )

    if any(term in q for term in ["market share", "participacao", "participa????o", "fabricantes", "marcas", "marca"]):
        return (
            "Market share Aquafast por fabricante",
            "SELECT * FROM ms_mercado_aquafast ORDER BY total_receita DESC LIMIT 20",
        )

    if any(term in q for term in ["estado", "uf", "vendas por estado"]):
        return (
            "Vendas Aquafast por estado",
            "SELECT * FROM vendas_caixas_estado ORDER BY receita_total DESC",
        )

    if any(term in q for term in ["cidade", "vendas por cidade"]):
        return (
            "Vendas Aquafast por cidade",
            "SELECT * FROM vendas_por_cidade ORDER BY receita_total DESC, caixas_vendidas DESC, cidade",
        )

    if any(term in q for term in ["rede", "bandeira", "ranking de redes"]):
        return (
            "Ranking de redes Aquafast",
            "SELECT * FROM ranking_redes ORDER BY total_receita DESC",
        )

    if any(term in q for term in ["categoria", "litragem", "mix", "produto por categoria"]):
        return (
            "Produtos por categoria Aquafast",
            "SELECT * FROM top_produtos_categoria ORDER BY caixas_vendidas DESC, produto_padrao LIMIT 50",
        )

    if (
        ("vendas" in q and "aquafast" in q and ("mes" in q or "mensal" in q))
        or any(
            term in q
            for term in [
                "vendas por mes",
                "vendas por m??s",
                "evolucao mensal",
                "evolu????o mensal",
                "receita por mes",
                "receita por m??s",
                "serie mensal",
                "s??rie mensal",
                "historico mensal",
                "hist??rico mensal",
                "comparativo mensal",
            ]
        )
    ):
        return "Vendas Aquafast por mes", "SELECT * FROM vendas_por_mes ORDER BY mes"

    if any(term in q for term in ["ultimos 12 meses", "??ltimos 12 meses", "ultimo ano", "??ltimo ano", "12 meses de vendas"]):
        return (
            "Ultimos 12 meses (vendas Aquafast por mes)",
            "SELECT * FROM vendas_por_mes ORDER BY mes DESC LIMIT 12",
        )

    if any(term in q for term in ["ultimos 6 meses", "??ltimos 6 meses", "6 meses de vendas"]):
        return (
            "Ultimos 6 meses (vendas Aquafast por mes)",
            "SELECT * FROM vendas_por_mes ORDER BY mes DESC LIMIT 6",
        )

    if any(term in q for term in ["potencial de venda", "mais potencial", "teriam mais potencial", "o que vender", "produto com potencial", "produtos com potencial", "distribuicao", "distribui????o"]):
        return (
            "Produtos Aquafast com maior potencial de venda",
            """
            SELECT
                produto_padrao,
                produto_original_exemplo,
                subgrupo_cigam,
                variacoes_produto_original,
                categoria,
                fabricante,
                marca,
                pdvs_com_venda,
                caixas_vendidas,
                total_receita,
                preco_medio_caixa
            FROM top_produtos_categoria
            ORDER BY pdvs_com_venda DESC, caixas_vendidas DESC, total_receita DESC, produto_padrao
            LIMIT 20
            """.strip(),
        )

    if any(term in q for term in ["pontos de venda", "ponto de venda", "pdv", "presentes hoje", "presente hoje", "presen??a hoje", "presenca hoje", "quantas lojas", "em quantos lojas", "lojas presentes"]):
        return (
            "Total de pontos de venda Aquafast",
            "SELECT COUNT(*) AS total_pontos_de_venda FROM ranking_clientes",
        )

    if any(term in q for term in ["quantos clientes", "quantas lojas", "quantos lojas", "em quantos lojas", "numero de clientes", "n??mero de clientes", "total de clientes", "quantos pdvs", "presente hoje"]):
        return (
            "Total de lojas Aquafast",
            "SELECT COUNT(*) AS total_lojas FROM ranking_clientes",
        )

    if any(term in q for term in ["quantos produtos", "numero de produtos", "n??mero de produtos", "total de produtos distintos", "quantos skus", "produtos aquafast"]):
        return (
            "Total de produtos Aquafast",
            "SELECT COUNT(DISTINCT produto_padrao) AS total_produtos FROM ranking_produtos",
        )

    if any(
        term in q
        for term in [
            "receita total",
            "faturamento total",
            "quanto vendemos no total",
            "valor total de vendas",
            "soma de todas as vendas",
        ]
    ) and not any(x in q for x in ["top", "ranking", "lista", "por cliente", "por produto"]):
        return (
            "Receita total agregada Aquafast",
            """
            SELECT
                ROUND(SUM(receita_total), 2) AS receita_total,
                CAST(SUM(caixas_vendidas) AS BIGINT) AS caixas_vendidas,
                ROUND(SUM(receita_total) / NULLIF(SUM(caixas_vendidas), 0), 2) AS ticket_medio_caixa
            FROM ranking_clientes
            """.strip(),
        )

    if any(term in q for term in ["ticket medio ponderado", "ticket m??dio ponderado", "ticket medio geral", "ticket m??dio geral"]):
        return (
            "Ticket medio por caixa Aquafast",
            """
            SELECT
                ROUND(SUM(receita_total) / NULLIF(SUM(caixas_vendidas), 0), 2) AS ticket_medio_caixa_ponderado,
                ROUND(AVG(ticket_medio_caixa), 2) AS ticket_medio_simples_entre_lojas
            FROM ranking_clientes
            """.strip(),
        )

    if any(term in q for term in ["clientes com mais pedidos", "mais pedidos por cliente", "quem mais compra em frequencia", "maior numero de pedidos"]):
        return (
            "Lojas com mais caixas (frequencia)",
            """
            SELECT cliente, unidades_scanntech, caixas_vendidas, receita_total, ticket_medio_caixa, primeira_compra, ultima_compra
            FROM ranking_clientes
            ORDER BY caixas_vendidas DESC, receita_total DESC
            LIMIT 50
            """.strip(),
        )

    if any(term in q for term in ["sem compra", "90 dias", "churn", "clientes sem compra", "lojas sem compra"]):
        return (
            "Lojas sem compra h?? mais de 90 dias",
            """
            SELECT cliente, ultima_compra, caixas_vendidas, receita_total
            FROM ranking_clientes
            WHERE ultima_compra < CURRENT_DATE - INTERVAL '90 days'
            ORDER BY receita_total DESC
            LIMIT 50
            """.strip(),
        )

    if any(term in q for term in ["1 compra", "uma compra", "apenas uma compra", "risco de churn", "clientes com apenas 1 compra"]):
        return (
            "Lojas com apenas 1 compra",
            """
            SELECT cliente, unidades_scanntech, caixas_vendidas, receita_total, primeira_compra, ultima_compra
            FROM ranking_clientes
            WHERE caixas_vendidas = 1
            ORDER BY receita_total DESC
            LIMIT 50
            """.strip(),
        )

    if any(term in q for term in ["resumo geral", "resumo do arquivo", "visao geral", "vis??o geral"]):
        return (
            "Resumo geral Aquafast",
            """
            SELECT
                COUNT(*) AS total_registros,
                COUNT(DISTINCT cliente) AS total_clientes,
                ROUND(SUM(receita_total), 2) AS receita_total,
                ROUND(AVG(ticket_medio_caixa), 2) AS ticket_medio_geral,
                MIN(primeira_compra) AS periodo_inicio,
                MAX(ultima_compra) AS periodo_fim
            FROM ranking_clientes
            """.strip(),
        )

    if re.match(r"^\s*(select|with|show|describe)\b", question, flags=re.IGNORECASE):
        return "Consulta SQL livre", question.strip().rstrip(";")

    raise ValueError(
        "Nao identifiquei uma consulta suportada. Use SQL direto ou use a IA do chat para gerar a consulta."
    )

def get_schema_snapshot() -> dict[str, Any]:
    con = open_connection()
    try:
        if CHAT_BACKEND == "mysql":
            tables = _execute_sql(
                con,
                f"""
                SELECT table_schema, table_name, table_type
                FROM information_schema.tables
                WHERE table_schema = '{MYSQL_DATABASE}'
                ORDER BY table_schema, table_name
                """,
            )[1]
            columns = _execute_sql(
                con,
                f"""
                SELECT table_schema, table_name, column_name, data_type, ordinal_position
                FROM information_schema.columns
                WHERE table_schema = '{MYSQL_DATABASE}'
                ORDER BY table_schema, table_name, ordinal_position
                """,
            )[1]
        else:
            tables = _execute_sql(
                con,
                """
                SELECT table_schema, table_name, table_type
                FROM information_schema.tables
                WHERE table_schema NOT IN ('information_schema', 'pg_catalog')
                ORDER BY table_schema, table_name
                """,
            )[1]
            columns = _execute_sql(
                con,
                """
                SELECT table_schema, table_name, column_name, data_type, ordinal_position
                FROM information_schema.columns
                WHERE table_schema NOT IN ('information_schema', 'pg_catalog')
                ORDER BY table_schema, table_name, ordinal_position
                """,
            )[1]
    finally:
        con.close()

    structured: list[dict[str, Any]] = []
    index: dict[tuple[str, str], dict[str, Any]] = {}
    for schema, table, table_type in tables:
        entry = {
            "schema": schema,
            "name": table,
            "type": table_type,
            "columns": [],
        }
        structured.append(entry)
        index[(schema, table)] = entry

    for schema, table, column_name, data_type, ordinal_position in columns:
        entry = index.get((schema, table))
        if entry is None:
            continue
        entry["columns"].append(
            {
                "name": column_name,
                "type": data_type,
                "position": ordinal_position,
            }
        )

    lines = []
    for entry in structured:
        column_text = ", ".join(f"{col['name']}:{col['type']}" for col in entry["columns"])
        lines.append(f"- {entry['name']} ({entry['type']}): {column_text}")

    return {
        "database": MYSQL_DATABASE if CHAT_BACKEND == "mysql" else DUCKDB_PATH.name,
        "table_count": sum(1 for item in structured if item["type"] == "BASE TABLE"),
        "view_count": sum(1 for item in structured if item["type"] == "VIEW"),
        "object_count": len(structured),
        "objects": structured,
        "summary_text": "\n".join(lines),
    }


def _get_schema_snapshot_light() -> dict[str, Any]:
    if CHAT_BACKEND == "mysql":
        con = mysql.connector.connect(**_mysql_config())
        try:
            tables = _execute_sql(
                con,
                f"""
                SELECT table_schema, table_name, table_type
                FROM information_schema.tables
                WHERE table_schema = '{MYSQL_DATABASE}'
                ORDER BY table_schema, table_name
                """,
            )[1]
            columns = _execute_sql(
                con,
                f"""
                SELECT table_schema, table_name, column_name, data_type, ordinal_position
                FROM information_schema.columns
                WHERE table_schema = '{MYSQL_DATABASE}'
                ORDER BY table_schema, table_name, ordinal_position
                """,
            )[1]
        finally:
            con.close()
    else:
        con = duckdb.connect(str(DUCKDB_PATH), read_only=True)
        try:
            tables = _execute_sql(
                con,
                """
                SELECT table_schema, table_name, table_type
                FROM information_schema.tables
                WHERE table_schema NOT IN ('information_schema', 'pg_catalog')
                ORDER BY table_schema, table_name
                """,
            )[1]
            columns = _execute_sql(
                con,
                """
                SELECT table_schema, table_name, column_name, data_type, ordinal_position
                FROM information_schema.columns
                WHERE table_schema NOT IN ('information_schema', 'pg_catalog')
                ORDER BY table_schema, table_name, ordinal_position
                """,
            )[1]
        finally:
            con.close()

    structured: list[dict[str, Any]] = []
    index: dict[tuple[str, str], dict[str, Any]] = {}
    for schema, table, table_type in tables:
        entry = {
            "schema": schema,
            "name": table,
            "type": table_type,
            "columns": [],
        }
        structured.append(entry)
        index[(schema, table)] = entry

    for schema, table, column_name, data_type, ordinal_position in columns:
        entry = index.get((schema, table))
        if entry is None:
            continue
        entry["columns"].append(
            {
                "name": column_name,
                "type": data_type,
                "position": ordinal_position,
            }
        )

    lines = []
    for entry in structured:
        column_text = ", ".join(f"{col['name']}:{col['type']}" for col in entry["columns"])
        lines.append(f"- {entry['name']} ({entry['type']}): {column_text}")

    return {
        "database": MYSQL_DATABASE if CHAT_BACKEND == "mysql" else DUCKDB_PATH.name,
        "table_count": sum(1 for item in structured if item["type"] == "BASE TABLE"),
        "view_count": sum(1 for item in structured if item["type"] == "VIEW"),
        "object_count": len(structured),
        "objects": structured,
        "summary_text": "\n".join(lines),
    }


@app.get("/health")
def health() -> dict[str, Any]:
    try:
        if CHAT_BACKEND == "mysql":
            con = mysql.connector.connect(**_mysql_config())
            try:
                total_registros = _execute_sql(con, "SELECT COUNT(*) AS total_registros FROM vta")[1][0][0]
            finally:
                con.close()
        else:
            con = duckdb.connect(str(DUCKDB_PATH), read_only=True)
            try:
                total_registros = _execute_sql(con, "SELECT COUNT(*) AS total_registros FROM scanntech")[1][0][0]
            finally:
                con.close()
        schema = _get_schema_snapshot_light()
        return {
            "ok": True,
            "db_path": MYSQL_DATABASE if CHAT_BACKEND == "mysql" else DUCKDB_PATH.name,
            "total_registros": total_registros,
            "tables": schema["table_count"],
            "views": schema["view_count"],
            "available_reports": AVAILABLE_REPORTS,
            "report_count": len(REPORT_SPECS),
            "schema_summary": schema["summary_text"],
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/schema")
def schema() -> dict[str, Any]:
    try:
        return {
            "ok": True,
            **get_schema_snapshot(),
            "available_reports": list_report_specs(),
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/reports")
def reports() -> dict[str, Any]:
    return {
        "ok": True,
        "report_count": len(REPORT_SPECS),
        "page_size_limit": REPORT_PAGE_SIZE_LIMIT,
        "reports": list_report_specs(),
    }


@app.get("/reports/{report_name}")
def report(report_name: str, page: int = 1, page_size: int = 50) -> dict[str, Any]:
    try:
        return {"ok": True, **run_report(report_name, page=page, page_size=page_size)}
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.get("/reports/{report_name}/export")
def report_export(report_name: str) -> dict[str, Any]:
    try:
        spec = get_report_spec(report_name)
        export_result = export_query(spec["sql"], spec["title"])
        return {
            "ok": True,
            "report_name": report_name,
            "title": spec["title"],
            "description": spec["description"],
            **export_result,
        }
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/query")
def query(request: SQLRequest) -> dict[str, Any]:
    try:
        sql = ensure_read_only_sql(request.sql)
        query_result = run_query(sql, row_cap=QUERY_RESULT_ROW_CAP)
        title = request.title.strip() or "Consulta SQL"
        route_meta = _route_metadata_for_response("", title, sql)
        return {
            "ok": True,
            "title": title,
            "question": "",
            "sql": sql,
            "source_note": _build_source_note_clean("", title, sql),
            **route_meta,
            **query_result,
        }
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/export")
def export(request: SQLRequest) -> dict[str, Any]:
    try:
        sql = ensure_read_only_sql(request.sql)
        export_result = export_query(sql, request.title.strip() or "Exportacao Excel")
        return {
            "ok": True,
            "title": request.title.strip() or "Exportacao Excel",
            "question": "",
            "sql": sql,
            "source_note": _build_source_note_clean("", request.title.strip() or "Exportacao Excel", sql),
            **export_result,
        }
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.get("/download/{file_name}")
def download(file_name: str) -> FileResponse:
    file_path = (EXPORT_DIR / file_name).resolve()
    export_root = EXPORT_DIR.resolve()
    if export_root not in file_path.parents or not file_path.exists():
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
    return FileResponse(
        path=file_path,
        filename=file_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/ask")
def ask(request: QuestionRequest) -> dict[str, Any]:
    try:
        title, sql = legacy_question_to_sql(request.question)
        route_meta = _route_metadata_for_response(request.question, title, sql)
        report_name = route_meta["report_name"] or _infer_report_name_from_sql(sql) or ""
        result = run_query(sql, row_cap=QUERY_RESULT_ROW_CAP)
        history_result = (
            _attach_history_metadata(
                result,
                pergunta=request.question,
                report_name=report_name,
                rows_returned=int(result.get("row_count", 0)),
            )
            if report_name
            else {}
        )
        return {
            "ok": True,
            "title": title,
            "question": request.question,
            "sql": sql,
            "source_note": _build_source_note_clean(request.question, title, sql),
            **route_meta,
            **history_result,
            **result,
        }
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.get("/official-questions")
def official_questions() -> dict[str, Any]:
    items = list_official_questions()
    return {"ok": True, "count": len(items), "items": items}


@app.get("/official-questions/{question_id}")
def official_question(question_id: str) -> dict[str, Any]:
    for route in OFFICIAL_QUESTION_ROUTES:
        if route.id == question_id:
            return {
                "ok": True,
                "id": repair_mojibake(route.id),
                "title": repair_mojibake(route.title),
                "description": repair_mojibake(route.description),
                "examples": [repair_mojibake(example) for example in route.examples],
                "source_note": repair_mojibake(route.source_note),
                "sql": repair_mojibake(route.sql),
            }
    raise HTTPException(status_code=404, detail="Pergunta oficial nao encontrada")


def main() -> None:
    parser = argparse.ArgumentParser(description="API local de analise Scanntech")
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--port", type=int, default=8000)
    args = parser.parse_args()
    uvicorn_run("api_fastapi:app", host=args.host, port=args.port, reload=False)


if __name__ == "__main__":
    main()
